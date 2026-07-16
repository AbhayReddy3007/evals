"""
IPD3 Validator — Independent Validation of Gemini Strategies

Unlike the orchestrator which compares Gemini vs Claude outputs,
this validator asks Claude to assess each Gemini strategy from FIRST
PRINCIPLES using pharmaceutical patent domain knowledge.

Claude does NOT see its own prior output. It validates Gemini's
strategies independently based on:
  - Patent law and claim construction principles
  - FDA 505(b)(2) regulatory pathway requirements
  - Known pharmaceutical design-around precedents
  - Scientific feasibility of proposed modifications

For each Gemini strategy, Claude returns:
  - valid / partially_valid / invalid
  - Reasoning grounded in domain knowledge
  - What would make it stronger (if partially valid)
  - Red flags (if invalid)

Usage:
    python3 valid.py                         # validate all drugs
    python3 valid.py --drug Semaglutide      # single drug
    python3 valid.py --skip-run              # use existing JSON only
"""

import os
import re
import sys
import json
import time
import random
import argparse
import subprocess
from datetime import datetime
from typing import Any, Dict, List, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
from anthropic import AnthropicVertex, RateLimitError

try:
    from dotenv import load_dotenv
    load_dotenv(override=True)
except ImportError:
    pass

try:
    from json_repair import repair_json as _repair_json_lib
except ImportError:
    _repair_json_lib = None

# ── Config ─────────────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

BQ_PROJECT_ID = os.getenv("BQ_PROJECT_ID", "cognito-prod-394707")
GCP_REGION    = os.getenv("GCP_REGION", "us-east5")
PROJECT_ID    = os.getenv("BQ_PROJECT_ID", "cognito-prod-394707")

CREDENTIALS_PATH = os.environ.get("CREDENTIALS_PATH") or os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "")
if CREDENTIALS_PATH and not os.path.isabs(CREDENTIALS_PATH):
    CREDENTIALS_PATH = os.path.join(SCRIPT_DIR, CREDENTIALS_PATH)

GCS_BUCKET    = os.getenv("GCS_BUCKET", "cognito-gcs")
GCS_BASE_PATH = os.getenv("GCS_EVAL_PATH", "Cognito_new/eval_reports")

CLAUDE_MODEL   = os.getenv("CLAUDE_MODEL", "claude-sonnet-4-6")
MAX_WORKERS    = 6
LLM_MAX_TOKENS = 16384

OUTPUT_DIR = os.getenv("IPD3_OUTPUT_DIR", os.path.join(SCRIPT_DIR, "ipd3_output"))


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_credentials():
    if CREDENTIALS_PATH and os.path.exists(CREDENTIALS_PATH):
        from google.oauth2 import service_account
        return service_account.Credentials.from_service_account_file(CREDENTIALS_PATH)
    return None


def get_claude_client() -> AnthropicVertex:
    return AnthropicVertex(region=GCP_REGION, project_id=PROJECT_ID)


def _call_claude(client: AnthropicVertex, prompt: str, max_retries: int = 3) -> Dict[str, Any]:
    for attempt in range(max_retries + 1):
        try:
            resp = client.messages.create(
                model=CLAUDE_MODEL, max_tokens=LLM_MAX_TOKENS,
                messages=[{"role": "user", "content": prompt}],
            )
            text = resp.content[0].text.strip()
            text = re.sub(r"^```(?:json)?", "", text).strip()
            text = re.sub(r"```$", "", text).strip()
            try:
                return json.loads(text)
            except json.JSONDecodeError:
                if _repair_json_lib:
                    return json.loads(_repair_json_lib(text))
                text = re.sub(r",\s*}", "}", text)
                text = re.sub(r",\s*]", "]", text)
                return json.loads(text)
        except RateLimitError:
            if attempt == max_retries: raise
            time.sleep(2.0 * (2 ** attempt) * (1 + random.uniform(0, 0.25)))
        except Exception as e:
            if attempt == max_retries:
                return {"error": str(e)}
            time.sleep(2.0 * (2 ** attempt))
    return {}


def _load_json_safe(path: str) -> List[Dict]:
    if not os.path.exists(path):
        print(f"[WARN] Not found: {path}")
        return []
    try:
        with open(path) as f:
            data = json.load(f)
        print(f"[JSON] Loaded {len(data)} rows from {path}")
        return data
    except Exception as e:
        print(f"[WARN] Failed: {path}: {e}")
        return []


def _find_script(name):
    p = os.path.join(SCRIPT_DIR, name)
    return p if os.path.exists(p) else name


# ── Load Gemini Strategies (grouped by category) ─────────────────────────────

def load_gemini_strategies(drug=None) -> List[Dict]:
    """Load Gemini's circumvention JSON and group strategies by drug+category."""
    rows = _load_json_safe(os.path.join(OUTPUT_DIR, "circumvention_gemini.json"))
    if not rows:
        return []

    grouped = {}
    for r in rows:
        key = (r.get("Drug_Name", "").strip(), r.get("Patent_Category", "").strip())
        if drug and key[0].lower() != drug.lower():
            continue
        if key not in grouped:
            grouped[key] = {
                "Drug_Name": key[0],
                "Patent_Category": key[1],
                "Patents": r.get("Patents", ""),
                "Overall_Difficulty": r.get("Overall_Difficulty", ""),
                "Key_Claim_Limitations": r.get("Key_Claim_Limitations", ""),
                "FDA_Precedents": r.get("FDA_Precedents", ""),
                "Summary": r.get("Summary", ""),
                "strategies": [],
            }
        strat = r.get("Strategy", "")
        if strat and strat != "No strategies identified":
            grouped[key]["strategies"].append({
                "strategy": strat,
                "rationale": r.get("Rationale", ""),
                "feasibility": r.get("Feasibility", ""),
                "regulatory_pathway": r.get("Regulatory_Pathway", ""),
                "prior_art_support": r.get("Prior_Art_Support", ""),
            })

    result = sorted(grouped.values(), key=lambda x: (x["Drug_Name"], x["Patent_Category"]))
    print(f"[GROUP] {len(result)} categories, {sum(len(r['strategies']) for r in result)} total strategies")
    return result


# ── Validation Prompt ─────────────────────────────────────────────────────────

VALIDATE_PROMPT = """
You are a senior pharmaceutical patent attorney with deep expertise in 505(b)(2)
regulatory strategy and patent design-around analysis.

You are asked to INDEPENDENTLY VALIDATE circumvention strategies produced by an LLM
(Gemini 2.5 Flash). You should evaluate each strategy purely on its merits using your
domain knowledge — NOT by comparing to any other system's output.

Drug: {drug}
Patent Category: {category}
Patents: {patents}
Claimed Overall Difficulty: {difficulty}
Key Claim Limitations (as identified by Gemini): {key_claims}

══ GEMINI STRATEGIES TO VALIDATE ══
{strategies_text}
══ END ══

For EACH strategy, assess:

1. **Technical Validity**: Is the proposed design-around technically sound?
   Does it correctly identify a claim limitation that can be omitted?

2. **Legal Soundness**: Would this approach actually avoid infringement under
   patent claim construction doctrine? Is the claim interpretation reasonable?

3. **Regulatory Feasibility**: Is this viable under a 505(b)(2) pathway?
   Would it require new clinical trials, or can it rely on the reference product?

4. **Scientific Plausibility**: Is the proposed modification scientifically feasible?
   Are there known examples of similar approaches in the pharmaceutical industry?

5. **Prior Art Check**: Does the cited prior art / FDA precedent actually exist and
   support the strategy? Flag any citations that appear fabricated.

6. **Red Flags**: Identify any hallucinated references, impossible chemistry,
   contradictory logic, or strategies that would INCREASE infringement risk.

Respond ONLY with valid JSON:
{{
  "drug": "{drug}",
  "category": "{category}",
  "difficulty_assessment_valid": <true if the claimed difficulty is reasonable, false otherwise>,
  "difficulty_reasoning": "<1 sentence on whether Easy/Moderate/Difficult is appropriate>",
  "key_claims_valid": <true if the identified claim limitations are plausible>,
  "strategies": [
    {{
      "strategy_text": "<the strategy being evaluated>",
      "verdict": "<valid | partially_valid | invalid>",
      "technical_validity": <1-5>,
      "legal_soundness": <1-5>,
      "regulatory_feasibility": <1-5>,
      "scientific_plausibility": <1-5>,
      "prior_art_verified": <true | false | unverifiable>,
      "red_flags": ["<any issues found>"],
      "strengths": ["<what this strategy gets right>"],
      "improvements_needed": ["<what would make it stronger, if partially_valid>"],
      "reasoning": "<2-3 sentence explanation of the verdict>"
    }}
  ],
  "overall_validity": "<valid | partially_valid | invalid>",
  "overall_reasoning": "<2-3 sentence summary: how reliable are Gemini's strategies for this category?>",
  "hallucination_count": <number of strategies with fabricated references or impossible claims>
}}
"""


# ── Validate One Category ─────────────────────────────────────────────────────

def validate_category(client, cat_data):
    strats = cat_data["strategies"]
    if not strats:
        return {
            "drug": cat_data["Drug_Name"],
            "category": cat_data["Patent_Category"],
            "strategies": [],
            "overall_validity": "N/A",
            "overall_reasoning": "No strategies to validate.",
            "hallucination_count": 0,
        }

    strat_lines = []
    for i, s in enumerate(strats, 1):
        strat_lines.append(f"Strategy {i}: {s['strategy']}")
        strat_lines.append(f"  Rationale: {s['rationale']}")
        strat_lines.append(f"  Feasibility: {s['feasibility']}")
        strat_lines.append(f"  Regulatory Pathway: {s['regulatory_pathway']}")
        strat_lines.append(f"  Prior Art Support: {s['prior_art_support']}")
        strat_lines.append("")

    prompt = VALIDATE_PROMPT.format(
        drug=cat_data["Drug_Name"],
        category=cat_data["Patent_Category"],
        patents=cat_data["Patents"],
        difficulty=cat_data["Overall_Difficulty"],
        key_claims=cat_data["Key_Claim_Limitations"],
        strategies_text="\n".join(strat_lines),
    )
    return _call_claude(client, prompt)


# ── GCS Upload ────────────────────────────────────────────────────────────────

def upload_to_gcs(local_path: str) -> str:
    from google.cloud import storage as gcs_storage
    credentials = _get_credentials()
    client = gcs_storage.Client(project=BQ_PROJECT_ID, credentials=credentials)
    bucket = client.bucket(GCS_BUCKET)
    blob_name = f"{GCS_BASE_PATH}/{os.path.basename(local_path)}"
    gcs_uri = f"gs://{GCS_BUCKET}/{blob_name}"
    print(f"[GCS] Uploading → {gcs_uri}")
    bucket.blob(blob_name).upload_from_filename(
        local_path,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return gcs_uri


# ── Excel Builder ─────────────────────────────────────────────────────────────

def save_validation_excel(results: List[Dict], drug=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = os.path.join(OUTPUT_DIR, f"ipd3_validation_{drug or 'all'}_{ts}.xlsx")
    wb = Workbook()

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    title_font = Font(name="Arial", bold=True, size=14, color="1F3864")
    sub_font = Font(name="Arial", bold=True, size=11, color="2F5496")
    cell_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    green = PatternFill("solid", fgColor="E2EFDA")
    red = PatternFill("solid", fgColor="FCE4EC")
    yellow = PatternFill("solid", fgColor="FFF9C4")
    gray = PatternFill("solid", fgColor="F5F5F5")

    def _vf(verdict):
        v = str(verdict).lower()
        if v == "valid": return green
        if v == "partially_valid": return yellow
        if v == "invalid": return red
        return None

    def _sf(val):
        try: v = int(float(val))
        except: return None
        return green if v >= 4 else yellow if v == 3 else red if v <= 2 else None

    def _hdr(ws, row, headers):
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=1+i, value=h)
            c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, center, thin

    def _row(ws, row, values, fills=None):
        for i, v in enumerate(values):
            c = ws.cell(row=row, column=1+i, value=v)
            c.font, c.alignment, c.border = cell_font, (left_al if i <= 1 else center), thin
            if fills and i < len(fills) and fills[i]:
                c.fill = fills[i]

    def _auto(ws, mn=10, mx=40):
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+2, mn), mx)

    # Flatten strategies for the detail sheet
    all_strats = []
    for r in results:
        val = r.get("validation", {})
        for s in val.get("strategies", []):
            all_strats.append({
                "Drug": r["Drug_Name"],
                "Category": r["Patent_Category"],
                **s,
            })

    total_strats = len(all_strats)
    valid_count = sum(1 for s in all_strats if str(s.get("verdict", "")).lower() == "valid")
    partial_count = sum(1 for s in all_strats if str(s.get("verdict", "")).lower() == "partially_valid")
    invalid_count = sum(1 for s in all_strats if str(s.get("verdict", "")).lower() == "invalid")
    total_halluc = sum(r.get("validation", {}).get("hallucination_count", 0) for r in results)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="IPD3 Validation — Gemini Strategies (independent review)").font = title_font
    ws.cell(row=2, column=1, value=f"Drug: {drug or 'All'}  |  Validator: {CLAUDE_MODEL}").font = sub_font
    ws.cell(row=3, column=1, value="Claude validates Gemini's strategies from first principles (no output comparison).").font = cell_font
    ws.merge_cells("A1:C1"); ws.merge_cells("A2:C2"); ws.merge_cells("A3:C3")

    _hdr(ws, 5, ["Metric", "Value"])
    summary_data = [
        ("Categories Evaluated", len(results)),
        ("Total Strategies", total_strats),
        ("", ""),
        ("Valid", f"{valid_count} ({round(valid_count/(total_strats or 1)*100,1)}%)"),
        ("Partially Valid", f"{partial_count} ({round(partial_count/(total_strats or 1)*100,1)}%)"),
        ("Invalid", f"{invalid_count} ({round(invalid_count/(total_strats or 1)*100,1)}%)"),
        ("", ""),
        ("Hallucinations Detected", total_halluc),
    ]
    for i, (m, v) in enumerate(summary_data):
        fill = None
        if "Valid" in str(m) and "Partially" not in str(m) and "Invalid" not in str(m): fill = green
        elif "Partially" in str(m): fill = yellow
        elif "Invalid" in str(m) or "Halluc" in str(m): fill = red
        _row(ws, 6+i, [m, v], fills=[gray, fill])
        if m: ws.cell(row=6+i, column=1).font = bold_font

    # Per-category summary
    cat_start = 6 + len(summary_data) + 2
    ws.cell(row=cat_start, column=1, value="Per-Category Verdict").font = sub_font
    _hdr(ws, cat_start+1, ["Drug", "Category", "Overall Validity", "Reasoning"])
    for i, r in enumerate(results):
        val = r.get("validation", {})
        v = val.get("overall_validity", "N/A")
        _row(ws, cat_start+2+i,
             [r["Drug_Name"], r["Patent_Category"], v, str(val.get("overall_reasoning", ""))],
             fills=[None, None, _vf(v), None])
    _auto(ws)
    ws.column_dimensions["D"].width = 60

    # ── Sheet 2: Strategy Detail ──────────────────────────────────────────
    ws2 = wb.create_sheet("Strategy Detail")
    _hdr(ws2, 1, [
        "Drug", "Category", "Strategy",
        "Verdict",
        "Technical", "Legal", "Regulatory", "Scientific",
        "Prior Art\nVerified", "Hallucinations",
        "Strengths", "Improvements Needed",
        "Reasoning",
    ])
    for i, s in enumerate(all_strats):
        rn = i + 2
        v = s.get("verdict", "")
        red_flags = "; ".join(s.get("red_flags", []))
        strengths = "; ".join(s.get("strengths", []))
        improvements = "; ".join(s.get("improvements_needed", []))
        pa = s.get("prior_art_verified")
        pa_str = "Yes" if pa is True else "No" if pa is False else "Unverifiable"

        vals = [
            s["Drug"], s["Category"], str(s.get("strategy_text", "")),
            v,
            s.get("technical_validity"), s.get("legal_soundness"),
            s.get("regulatory_feasibility"), s.get("scientific_plausibility"),
            pa_str, red_flags,
            strengths, improvements,
            str(s.get("reasoning", "")),
        ]
        fills = [
            None, None, None,
            _vf(v),
            _sf(s.get("technical_validity")), _sf(s.get("legal_soundness")),
            _sf(s.get("regulatory_feasibility")), _sf(s.get("scientific_plausibility")),
            green if pa is True else (red if pa is False else yellow),
            red if red_flags else None,
            None, None, None,
        ]
        _row(ws2, rn, vals, fills=fills)

    _auto(ws2, mn=8, mx=30)
    ws2.column_dimensions["C"].width = 55
    ws2.column_dimensions["J"].width = 40
    ws2.column_dimensions["K"].width = 40
    ws2.column_dimensions["L"].width = 40
    ws2.column_dimensions["M"].width = 60

    wb.save(fname)
    print(f"[EXCEL] → {fname}")
    try:
        upload_to_gcs(fname)
    except Exception as e:
        print(f"[WARN] GCS upload failed: {e}")
    return fname


# ── Main ──────────────────────────────────────────────────────────────────────

def validate(drug=None, skip_run=False, extra_args=None, no_bq=False):
    start = time.time()

    print("=" * 70)
    print("IPD3 VALIDATOR — Independent Validation of Gemini Strategies")
    print(f"  Drug: {drug or 'ALL'}  |  Validator: {CLAUDE_MODEL}")
    print(f"  Mode: First-principles review (no output comparison)")
    print("=" * 70)

    # Optionally run the Gemini pipeline first
    if not skip_run:
        script = _find_script("ipd3bq__6_.py")
        cmd = [sys.executable, script]
        if drug: cmd.append(drug)
        if extra_args: cmd.extend(extra_args)
        if no_bq: cmd.append("--no-bq")
        print(f"\n  Running Gemini pipeline: {' '.join(cmd)}")
        subprocess.run(cmd)

    # Load Gemini strategies
    categories = load_gemini_strategies(drug)
    if not categories:
        print("[ERROR] No Gemini strategies found.")
        return {"error": "No data"}

    print(f"\nValidating {len(categories)} categories, "
          f"{sum(len(c['strategies']) for c in categories)} strategies ...")

    client = get_claude_client()
    results = []

    def _validate_one(cat):
        d, c = cat["Drug_Name"], cat["Patent_Category"]
        n = len(cat["strategies"])
        print(f"  [{d}/{c}] {n} strategies ...")
        val = validate_category(client, cat)
        return {**cat, "validation": val}

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(_validate_one, c): c for c in categories}
        for f in as_completed(futures):
            results.append(f.result())

    results.sort(key=lambda r: (r["Drug_Name"], r["Patent_Category"]))

    # Summary
    all_v = []
    for r in results:
        for s in r.get("validation", {}).get("strategies", []):
            all_v.append(s.get("verdict", ""))

    valid = sum(1 for v in all_v if v == "valid")
    partial = sum(1 for v in all_v if v == "partially_valid")
    invalid = sum(1 for v in all_v if v == "invalid")
    total = len(all_v) or 1

    print(f"\n{'='*60}")
    print(f"VALIDATION SUMMARY")
    print(f"  Strategies: {len(all_v)} total")
    print(f"    Valid           : {valid} ({round(valid/total*100,1)}%)")
    print(f"    Partially valid : {partial} ({round(partial/total*100,1)}%)")
    print(f"    Invalid         : {invalid} ({round(invalid/total*100,1)}%)")
    print(f"  Elapsed: {time.time()-start:.1f}s")
    print("=" * 60)

    save_validation_excel(results, drug)
    return {"results": results}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="IPD3 Validator — Independent Strategy Validation")
    parser.add_argument("--drug", default=None)
    parser.add_argument("--skip-run", action="store_true", help="Skip Gemini pipeline, use existing JSON")
    parser.add_argument("--no-bq", action="store_true")
    parser.add_argument("--rerun", action="store_true")
    parser.add_argument("--csv-input", default=None)
    args = parser.parse_args()

    extra = []
    if args.rerun: extra.append("--rerun")
    if args.csv_input: extra.extend(["--csv-input", args.csv_input])

    validate(drug=args.drug, skip_run=args.skip_run, extra_args=extra or None, no_bq=args.no_bq)
