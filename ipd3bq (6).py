"""
Patent Thicket Analysis & Circumvention Strategy (505(b)(2))

Processes NON-BLOCKING patents per drug:
  Section 1: Patent list
  Section 2: Category count table (patent thicket)
  Section 3: Density & Diversity interpretations
  Section 4: Circumvention Analysis per Patent Category
             (uses AlloyDB patent chunks + Gemini + Google Search grounding)

Changes from original:
  - fetch_relevant_chunks: no longer scans all collections when a patent is
    absent from the primary collection. Returns [] immediately.
  - _analyse_one_category: when no AlloyDB chunks are found for a category,
    runs a Google Search-grounded Gemini call to produce a full circumvention
    narrative (same JSON shape as the normal path) instead of a bare placeholder.
"""

import os
import re
import json
import time
import asyncio
import argparse
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
# google.generativeai is deprecated. Use google.genai (new SDK) only.
from google import genai
from google.genai import types
from google.cloud import bigquery
from google.oauth2 import service_account
try:
    from dotenv import load_dotenv
    load_dotenv(override=True)
except ImportError:
    pass  # Not needed on Cloud Run

# ── Import AlloyDB client ─────────────────────────────────────────────────────
# `alloydb_client.py` lives in `cog/` (a sibling package). Depending on how
# this script is deployed it may sit at any of:
#   <script_dir>/cog/alloydb_client.py            (running from pipeline/)
#   <script_dir>/../cog/alloydb_client.py         (container layout: /app/Pipeline/ + /app/cog/)
#   <script_dir>/alloydb_client.py                (rare — flat layout)
import sys
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

_alloydb_candidates = [
    os.path.join(SCRIPT_DIR, "cog"),
    os.path.join(SCRIPT_DIR, os.pardir, "cog"),
    SCRIPT_DIR,
]
_alloydb_dir = next(
    (d for d in _alloydb_candidates
     if os.path.exists(os.path.join(d, "alloydb_client.py"))),
    None,
)
if _alloydb_dir is None:
    raise ModuleNotFoundError(
        "alloydb_client.py not found. Looked in:\n  "
        + "\n  ".join(os.path.abspath(d) for d in _alloydb_candidates)
        + "\nMake sure cog/alloydb_client.py is included in your container "
          "image and lives next to (or as a sibling of) this script."
    )
_alloydb_dir = os.path.abspath(_alloydb_dir)
if _alloydb_dir not in sys.path:
    sys.path.insert(0, _alloydb_dir)
print(f"[IPD3BQ] alloydb_client located at: {_alloydb_dir}/alloydb_client.py")

from alloydb_client import AlloyDBClient

# Imported as `glp1_universe` (not `drug_filter`) because this file already
# uses the local name `drug_filter` as a parameter/CLI concept — a single
# drug name to restrict processing to. See process_patents()'s signature.
from cog import drug_filter as glp1_universe

# ── BigQuery Config (from .env) ───────────────────────────────────────────────
BQ_PROJECT_ID  = os.getenv("BQ_PROJECT_ID", "cognito-prod-394707")
BQ_DATASET_ID  = os.getenv("BQ_DATASET_ID", "cognito_prod_datamart")
BQ_TABLE_ID    = os.getenv("BQ_LOE_TABLE_NAME", "Master_LOE")
CREDENTIALS_PATH = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "")
BQ_LOCATION    = os.getenv("BQ_LOCATION", "asia-south1")


def load_data_from_bigquery() -> pd.DataFrame:
    """
    Loads patent data from BigQuery into a DataFrame.
    All columns arrive as STR; numeric coercion happens downstream.
    Uses GOOGLE_APPLICATION_CREDENTIALS for authentication if provided, else ADC.

    NOTE: BQ_LOCATION must match the region where the dataset lives
    (e.g. 'asia-south1'). Omitting it causes jobs to be routed to
    the default US region and authentication fails with Invalid JWT Signature.
    """
    table_ref = f"{BQ_PROJECT_ID}.{BQ_DATASET_ID}.{BQ_TABLE_ID}"
    print(f"[BigQuery] Connecting to: {table_ref}  (location={BQ_LOCATION})")

    credentials = _get_credentials()
    client = bigquery.Client(project=BQ_PROJECT_ID, credentials=credentials)

    query = f"""
    SELECT * EXCEPT(rn) FROM (
        SELECT *, ROW_NUMBER() OVER (
            PARTITION BY Patent_Number
            ORDER BY created_at DESC
        ) AS rn
        FROM `{table_ref}`
    ) WHERE rn = 1
    """
    print(f"[BigQuery] Running ROW_NUMBER dedup query on: {table_ref}")
    df = client.query(query, location=BQ_LOCATION).to_dataframe()
    print(f"[BigQuery] Loaded {len(df)} rows, columns: {list(df.columns)}")

    # BQ columns use underscores; rename to match the rest of the code (spaces)
    bq_to_code = {
        "Drug_Name":                              "Drug Name",
        "Patent_Number":                          "Patent Number",
        "Jurisdiction":                           "Jurisdiction",
        "Tag":                                    "Tag",
        "Blocking_Category":                      "Blocking Category",
        "Reason":                                 "Reason",
        "Step_1_Claim_Category":                  "Step 1 Claim Category",
        "Step_2_Matched_Elements":                "Step 2 Matched Elements",
        "S2_Active_Ingredient__Form":             "S2 Active Ingredient Form",
        "S2_Formulation_Details":                 "S2 Formulation Details",
        "S2_Route_of_Administration":             "S2 Route of Administration",
        "S2_Device_Description":                  "S2 Device Description",
        "S2_Combination_TechProcess":             "S2 Combination TechProcess",
        "Step_3_Technical_Barrier":               "Step 3 Technical Barrier",
        "Step_3_Confidence":                      "Step 3 Confidence",
        "Step_3_Evidence_Type":                   "Step 3 Evidence Type",
        "Step_3_Evidence_Summary":                "Step 3 Evidence Summary",
        "Step_4_Blocking_Indicator":              "Step 4 Blocking Indicator",
        "Step_4_Confidence":                      "Step 4 Confidence",
        "Step_4_Regulatory_Failure_if_Removed":   "Step 4 Regulatory Failure if Removed",
        "Step_4_Bridging_Studies_Required":       "Step 4 Bridging Studies Required",
        "Step_4_Formulation_Consistent_Across_Phases": "Step 4 Formulation Consistent Across Phases",
        "Step_4_Reason":                          "Step 4 Reason",
        "Step_5_Novel__Difficult":                "Step 5 Novel Difficult",
        "Step_5_Novelty_Signal":                  "Step 5 Novelty Signal",
        "Step_5_FirstinClass":                    "Step 5 FirstinClass",
        "Step_5_Prior_Failed_Attempts":           "Step 5 Prior Failed Attempts",
        "Step_5_Complex_Implementation":          "Step 5 Complex Implementation",
        "Step_5_Confidence":                      "Step 5 Confidence",
        "Step_5_Reason":                          "Step 5 Reason",
        "Filing_Date":                            "Filing Date",
        "Grant_Date":                             "Grant Date",
        "PTE_months":                             "PTE months",
        "Pediatric_Exclusivity":                  "Pediatric Exclusivity",
        "Phase":                                  "Phase",
        "Launch_Date":                            "Launch Date",
        "Approval_Date":                          "Approval Date",
        "Approval_Date_Source":                   "Approval Date Source",
        "Est_Approval_Year":                      "Est Approval Year",
        "Exclusivity_Year":                       "Exclusivity Year",
        "Controlling_Patent_Expiry_Year":         "Controlling Patent Expiry Year",
        "Years_to_Entry":                         "Years to Entry",
        "Avg_Years_to_Entry":                     "Avg Years to Entry",
        "Score":                                  "Score",
        "Avg_Years_to_Entry_US__EP":              "Avg Years to Entry US EP",
        "IP_Dimension_1_Score":                   "IP Dimension 1 Score",
        "Source_File":                            "Source File",
        "Type":                                   "Type",
        "No_Of_Forecasted_Patents":               "No Of Forecasted Patents",
    }
    df.rename(columns=bq_to_code, inplace=True)
    print(f"[BigQuery] Columns after rename: {list(df.columns)}")
    return df

# ── Config ────────────────────────────────────────────────────────────────────
GEMINI_MODEL = "gemini-2.5-flash"
API_KEY      = os.environ.get("GEMINI_API_KEY", "")

# ── Styles ────────────────────────────────────────────────────────────────────
HDR_FONT      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HDR_FILL      = PatternFill("solid", start_color="2F5496")
DRUG_FONT     = Font(name="Arial", bold=True, size=13, color="1F3864")
DRUG_FILL     = PatternFill("solid", start_color="BDD7EE")
CELL_FONT     = Font(name="Arial", size=10)
BOLD_FONT     = Font(name="Arial", bold=True, size=10)
ALT_FILL      = PatternFill("solid", start_color="DCE6F1")
GREEN_FILL    = PatternFill("solid", start_color="E2EFDA")
YELLOW_FILL   = PatternFill("solid", start_color="FFEB9C")
ORANGE_FILL   = PatternFill("solid", start_color="F4B942")
RED_FILL      = PatternFill("solid", start_color="FF6B6B")

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Circumvention Prompts ─────────────────────────────────────────────────────

CIRCUMVENTION_PROMPT = """
You are a pharmaceutical patent attorney specializing in 505(b)(2) regulatory strategy
and patent design-around analysis.

Below are the most relevant excerpts from non-blocking patent(s) in the "{claim_category}"
category for the drug "{drug}":

--- PATENT EXCERPTS ---
{chunks}
--- END EXCERPTS ---

Patents in this category: {patent_numbers}

Drug: {drug}

Objective: Identify potential non-infringing product designs that avoid the innovator's
patent claims. A design-around avoids infringement by omitting at least one key claim
limitation from each relevant patent.

Evaluation Steps:
1. Identify the key claim limitations in the independent claims for each patent.
2. For the "{claim_category}" category, evaluate whether alternative solutions exist
   that omit at least one key limitation.
3. Consider known approaches from FDA Drugs@FDA approval packages, FDA Orange Book
   listed patents, and scientific literature.
4. Assess whether a non-infringing product modification could still be regulatorily
   viable under a 505(b)(2) pathway.

For context, here are example design-around strategies by category:
- Formulation: Alternative excipient or stabilization chemistry; device-based mixing prior to injection
- API Ratio / Dosing: Different clinically justified ratio outside scope of dosing claims
- Delivery Device: Mechanically distinct injector design avoiding utility or design patent claims
- Manufacturing: Different synthesis pathway (process patents are often narrow)
- Method of Use: New therapeutic indication (may obtain separate exclusivity)

Respond ONLY with a valid JSON object — no markdown, no explanation outside the JSON:
{{
  "claim_category": "{claim_category}",
  "key_claim_limitations": ["<limitation 1>", "<limitation 2>", "<limitation 3>"],
  "design_around_strategies": [
    {{
      "strategy": "<concise description of the design-around approach>",
      "rationale": "<why this avoids infringement — which claim limitation is omitted>",
      "feasibility": "High / Medium / Low",
      "regulatory_pathway": "<how this could work under 505(b)(2) or other pathway>",
      "prior_art_support": "<any known precedent from FDA approvals, Orange Book, or literature>"
    }}
  ],
  "white_space_opportunities": ["<opportunity 1>", "<opportunity 2>"],
  "overall_circumvention_difficulty": "Easy / Moderate / Difficult",
  "summary": "<2-3 sentence overall assessment of circumvention potential for this category>"
}}
"""

CIRCUMVENTION_SEARCH_PROMPT = """
You are a pharmaceutical patent analyst. Search for 505(b)(2) design-around opportunities
for the drug "{drug}" in the "{claim_category}" patent category.

Specifically search for:
1. FDA Drugs@FDA: alternative formulations, dosing regimens, delivery devices, or new indications
   for {drug} or similar molecules
2. FDA Orange Book: listed patents and expiry dates to identify product features NOT protected
3. Scientific literature: {drug} + {claim_category_lower} alternatives, stability studies,
   delivery innovations

Return JSON:
{{
  "fda_precedents": ["<relevant FDA approval or product that demonstrates an alternative approach>"],
  "orange_book_gaps": ["<product feature or design element not covered by listed patents>"],
  "literature_alternatives": ["<published alternative approach from scientific literature>"],
  "regulatory_viability": "<assessment of whether non-infringing modifications are regulatorily feasible>"
}}

If no information found for a field, return an empty list for that field.
"""

# ── Online Circumvention Prompt (used when no AlloyDB chunks are available) ───
# Combines the chunk-based and search-based prompts into a single grounded call
# so we still produce a full, consistent result when local patent text is absent.

CIRCUMVENTION_ONLINE_PROMPT = """
You are a pharmaceutical patent attorney specialising in 505(b)(2) regulatory
strategy and patent design-around analysis.

No patent text is available in the local database for the "{claim_category}"
category patents of "{drug}" (patent numbers: {patent_numbers}).

Using publicly available information — FDA Drugs@FDA, the FDA Orange Book,
scientific literature, and patent databases (e.g. USPTO, Espacenet) — perform
a full circumvention analysis for this category.

Evaluation Steps:
1. Identify the typical key claim limitations found in "{claim_category}" patents
   for drugs like "{drug}" based on public patent databases.
2. Evaluate whether alternative solutions exist that omit at least one key limitation.
3. Consider known approaches from FDA Drugs@FDA approval packages, FDA Orange Book
   listed patents, and scientific literature.
4. Assess whether a non-infringing product modification could still be regulatorily
   viable under a 505(b)(2) pathway.

Return ONLY a valid JSON object — no markdown, no explanation outside the JSON:
{{
  "claim_category": "{claim_category}",
  "key_claim_limitations": ["<limitation 1>", "<limitation 2>", "<limitation 3>"],
  "design_around_strategies": [
    {{
      "strategy": "<concise description of the design-around approach>",
      "rationale": "<why this avoids infringement — which claim limitation is omitted>",
      "feasibility": "High / Medium / Low",
      "regulatory_pathway": "<how this could work under 505(b)(2) or other pathway>",
      "prior_art_support": "<any known precedent from FDA approvals, Orange Book, or literature>"
    }}
  ],
  "white_space_opportunities": ["<opportunity 1>", "<opportunity 2>"],
  "overall_circumvention_difficulty": "Easy / Moderate / Difficult",
  "fda_precedents": ["<relevant FDA approval or product demonstrating an alternative approach>"],
  "orange_book_gaps": ["<product feature or design element not covered by listed patents>"],
  "literature_alternatives": ["<published alternative approach from scientific literature>"],
  "regulatory_viability": "<assessment of whether non-infringing modifications are regulatorily feasible>",
  "summary": "<2-3 sentence overall assessment of circumvention potential for this category, noting that analysis is based on public sources as patent text was unavailable in the local database>"
}}
"""

# ── Thicket Score ─────────────────────────────────────────────────────────────
HIGH_DOMAIN_CATEGORIES = {
    "chemistry", "device", "manufacturing", "software",
    "api", "formulation", "delivery device", "method of use",
}

THICKET_SCORE_LABELS = {
    5: "Exceptional – No meaningful secondary patent fence",
    4: "Strong – High probability of design-around",
    3: "Moderate – Requires structured strategy",
    2: "Weak – Circumvention and litigation become costly",
    1: "Poor – Dense Patent Thicket",
}

THICKET_SCORE_FILLS = {
    5: PatternFill("solid", start_color="C6EFCE"),
    4: PatternFill("solid", start_color="92D050"),
    3: PatternFill("solid", start_color="FFEB9C"),
    2: PatternFill("solid", start_color="FFC7CE"),
    1: PatternFill("solid", start_color="FF0000"),
}


def compute_diversity_score(active_areas: int) -> int:
    """
    No. of Technology Areas | Diversity Score
    1–2                     | 5
    3                       | 3
    4–5                     | 2
    ≥6                      | 1
    0                       | 5  (no patents at all)
    """
    if active_areas <= 2:
        return 5
    if active_areas == 3:
        return 3
    if active_areas <= 5:
        return 2
    return 1


def compute_density_score(adjusted_count: float) -> int:
    """
    Total Adjusted Count | Density Score
    0–5                  | 5
    6–15                 | 3
    16–30                | 2
    >30                  | 1
    """
    if adjusted_count <= 5:
        return 5
    if adjusted_count <= 15:
        return 3
    if adjusted_count <= 30:
        return 2
    return 1


def compute_final_score(total_patents: int, combined_total: int,
                        adjusted_count: float, active_areas: int) -> dict:
    """
    Computes diversity score, density score, base score, validation %,
    and final score for a jurisdiction block.

    If diversity_score == 5, adjusted_count is reduced by 1 before
    computing the density score.

    Base Score = min(density_score, diversity_score)
    If density_score <= 2 AND diversity_score <= 2: base_score -= 1
    If validation_pct < 50: base_score += 1
    Final Score = max(1, min(5, base_score))
    """
    diversity_score = compute_diversity_score(active_areas)

    # If diversity score is 5 (low diversity), reduce adjusted count by 1
    effective_adjusted = adjusted_count - 1 if diversity_score == 5 else adjusted_count
    effective_adjusted = max(0, effective_adjusted)  # don't go negative

    density_score = compute_density_score(effective_adjusted)

    base_score = min(density_score, diversity_score)

    if density_score <= 2 and diversity_score <= 2:
        base_score -= 1

    # Validation % = (Total Number Of Patents / Total Combined) * 100
    if combined_total > 0:
        validation_pct = (total_patents / combined_total) * 100
    else:
        validation_pct = 100.0  # no patents at all

    if validation_pct < 50:
        base_score += 1

    final_score = max(1, min(5, base_score))

    return {
        "diversity_score": diversity_score,
        "density_score": density_score,
        "effective_adjusted_count": effective_adjusted,
        "base_score": base_score,
        "validation_pct": round(validation_pct, 1),
        "final_score": final_score,
        "final_label": THICKET_SCORE_LABELS[final_score],
    }


# ── Thicket thresholds ────────────────────────────────────────────────────────
def patent_count_label(n):
    if n == 0:   return "No Patents"
    if n <= 5:   return "Low Density Patent Thicket (0–5)"
    if n <= 15:  return "Moderate Density Patent Thicket (6–15)"
    if n <= 30:  return "High Density Patent Thicket (16–30)"
    return           "Dense / Multi-Layer Patent Thicket (>30)"

def patent_count_fill(n):
    if n == 0:   return PatternFill("solid", start_color="F2F2F2")
    if n <= 5:   return GREEN_FILL
    if n <= 15:  return YELLOW_FILL
    if n <= 30:  return ORANGE_FILL
    return           RED_FILL

def diversity_label(areas):
    if areas == 0:  return "No Active Areas"
    if areas <= 2:  return "Low Diversity (1–2 areas)"
    if areas == 3:  return "Moderate Diversity (3 areas)"
    if areas <= 5:  return "High Diversity (4–5 areas)"
    return              "Dense / Multi-Domain Patent Thicket (>5 areas)"

def diversity_fill(areas):
    if areas == 0:  return PatternFill("solid", start_color="F2F2F2")
    if areas <= 2:  return GREEN_FILL
    if areas == 3:  return YELLOW_FILL
    if areas <= 5:  return ORANGE_FILL
    return              RED_FILL

# ── Excel Helpers ─────────────────────────────────────────────────────────────
def write_cell(ws, row, col, value, font=None, fill=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = font      or CELL_FONT
    cell.alignment = alignment or CENTER
    cell.border    = THIN
    if fill:
        cell.fill = fill
    return cell

def write_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        write_cell(ws, row, col_start + i, h, font=HDR_FONT, fill=HDR_FILL)

def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def merge_title(ws, row, col1, col2, value, font=None, fill=None):
    ws.merge_cells(f"{col1}{row}:{col2}{row}")
    c = ws.cell(row=row, column=ord(col1) - 64, value=value)
    c.font, c.alignment, c.border = font or DRUG_FONT, CENTER, THIN
    if fill:
        c.fill = fill
    for col in range(ord(col1) - 64, ord(col2) - 63):
        ws.cell(row=row, column=col).border = THIN

# ── AlloyDB Helpers ───────────────────────────────────────────────────────────

_alloydb_client = AlloyDBClient()

def get_chroma_clients():
    """Return a list containing the AlloyDB client (backward-compatible name)."""
    return [_alloydb_client]

# Keep backward-compatible single-client helper
def get_chroma_client():
    return _alloydb_client

def collection_name(drug: str) -> str:
    return f"patents_{drug.strip().replace(' ', '_')}"

def fetch_relevant_chunks(client, drug: str, source_file: str,
                          sections: list[str], top_k: int = 12) -> list[str]:
    """
    Fetch chunks from AlloyDB using the 'Source File' column value,
    which matches the 'filename' metadata field stored during ingestion.

    Only the drug-specific collection (patents_{drug}) is searched.
    If the patent is not found there, an empty list is returned immediately —
    no fallback scan of all collections is performed.
    """
    if not source_file.strip():
        print(f"    [WARN] Empty source_file passed to fetch_relevant_chunks")
        return []

    exact_filename = source_file.strip()

    def _query_collection(collection) -> list[str]:
        """Return all matching docs from a single collection."""
        docs = []

        # Phase 1: compound filter
        try:
            results = collection.get(
                where={"$and": [
                    {"filename": {"$eq": exact_filename}},
                    {"chunk_index": {"$gte": 0}},
                ]},
                include=["documents", "metadatas"],
            )
            docs = results.get("documents", [])
        except Exception:
            pass

        # Phase 2: filename-only filter (fallback within the same collection)
        if not docs:
            try:
                results = collection.get(
                    where={"filename": {"$eq": exact_filename}},
                    include=["documents", "metadatas"],
                )
                docs = [
                    d for d, m in zip(
                        results.get("documents", []),
                        results.get("metadatas", []),
                    )
                    if m.get("chunk_index", -1) >= 0
                ]
            except Exception as e:
                print(f"    [WARN] Query failed on '{collection.name}': {e}")

        return docs

    def _rank_and_slice(docs: list[str]) -> list[str]:
        def relevance_score(text: str) -> int:
            t = text.lower()
            return sum(1 for kw in sections if kw in t)
        return sorted(docs, key=relevance_score, reverse=True)[:top_k]

    # Primary collection only — no fallback scan of all collections
    primary_coll_name = collection_name(drug)
    try:
        primary_coll = _alloydb_client.get_collection(primary_coll_name)
        docs = _query_collection(primary_coll)
        if docs:
            print(f"    [INFO] Found {len(docs)} chunks in collection "
                  f"'{primary_coll_name}' for '{exact_filename}'")
            return _rank_and_slice(docs)
    except Exception as e:
        print(f"    [WARN] Could not query collection '{primary_coll_name}': {e}")

    print(f"    [INFO] '{exact_filename}' not found in primary collection "
          f"'{primary_coll_name}' — skipping patent (no fallback scan).")
    return []

# ── Gemini Helpers ────────────────────────────────────────────────────────────

_genai_client = None

def _get_genai_client():
    global _genai_client
    if _genai_client is None:
        _genai_client = genai.Client(api_key=API_KEY)
    return _genai_client


def call_gemini(prompt: str) -> dict:
    client = _get_genai_client()
    response = client.models.generate_content(
        model=GEMINI_MODEL,
        contents=prompt,
    )
    text = (response.text or "").strip()
    text = re.sub(r"^```(?:json)?", "", text).strip()
    text = re.sub(r"```$", "", text).strip()
    return json.loads(text)

def _extract_json_from_response(text: str) -> str | None:
    if not text:
        return None
    text = re.sub(r"^```(?:json)?", "", text.strip()).strip()
    text = re.sub(r"```$", "", text).strip()
    match = re.search(r"\{.*\}", text, re.DOTALL)
    return match.group(0) if match else None

def _repair_json(text: str):
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        text = re.sub(r",\s*}", "}", text)
        text = re.sub(r",\s*]", "]", text)
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return text

# ── Circumvention Analysis ────────────────────────────────────────────────────

import concurrent.futures

# Thread pool for running blocking Gemini calls concurrently
_executor = concurrent.futures.ThreadPoolExecutor(max_workers=16)

# Concurrency limit: how many categories/patents analysed at once
CATEGORY_CONCURRENCY = 8


async def _analyse_one_category(
    drug_name: str,
    category: str,
    patents: list[dict],
    chroma_client,
    gemini_client,
    search_config,
) -> tuple[str, dict]:
    """Analyse a single category — runs Gemini scoring + FDA search.

    If no AlloyDB chunks are found for any patent in this category, falls
    back to a Google Search-grounded Gemini call (CIRCUMVENTION_ONLINE_PROMPT)
    that populates the same JSON fields so downstream consumers always receive
    a consistent result. The fallback does NOT run a second supplementary
    search call — the online prompt already requests all required fields.
    """
    print(f"\n[Circumvention] Category: {category} ({len(patents)} patent(s))")

    patent_numbers = [p["patent_number"] for p in patents]

    all_chunks = []
    for p in patents:
        chunks = fetch_relevant_chunks(
            chroma_client,
            p["drug_name"],
            p["source_file"],
            sections=["independent claim", "claim 1", "claims", "background",
                      "summary of invention", "abstract", "example",
                      "formulation", "device", "method", "dosing", "manufacturing"],
            top_k=8,
        )
        if chunks:
            all_chunks.extend(chunks[:6])

    loop = asyncio.get_event_loop()

    if all_chunks:
        # ── Normal path: use patent chunks ───────────────────────────────────
        chunks_text = "\n\n---\n\n".join(
            f"[Chunk {i+1}]\n{c}" for i, c in enumerate(all_chunks[:15])
        )
        prompt = CIRCUMVENTION_PROMPT.format(
            claim_category=category,
            drug=drug_name,
            chunks=chunks_text,
            patent_numbers=", ".join(str(x) for x in patent_numbers),
        )
        try:
            result = await loop.run_in_executor(_executor, call_gemini, prompt)
        except Exception as e:
            print(f"    [WARN] Gemini scoring error: {e}")
            result = {
                "claim_category": category,
                "design_around_strategies": [],
                "summary": f"Error: {e}",
            }

        # Supplement with FDA/literature search
        print(f"    Searching FDA/literature for {category} alternatives...")
        search_prompt = CIRCUMVENTION_SEARCH_PROMPT.format(
            drug=drug_name,
            claim_category=category,
            claim_category_lower=category.lower(),
        )
        try:
            search_resp = await gemini_client.aio.models.generate_content(
                model="gemini-2.5-flash", contents=search_prompt, config=search_config
            )
            search_json = _extract_json_from_response(
                search_resp.text.strip() if search_resp.text else ""
            )
            if search_json:
                search_data = _repair_json(search_json)
                if isinstance(search_data, str):
                    search_data = json.loads(search_data)
                result["fda_precedents"]         = search_data.get("fda_precedents", [])
                result["orange_book_gaps"]        = search_data.get("orange_book_gaps", [])
                result["literature_alternatives"] = search_data.get("literature_alternatives", [])
                result["regulatory_viability"]    = search_data.get("regulatory_viability", "")
        except Exception as e:
            print(f"    [WARN] FDA/literature search error: {e}")
            result["fda_precedents"]         = []
            result["orange_book_gaps"]        = []
            result["literature_alternatives"] = []
            result["regulatory_viability"]    = f"Search error: {e}"

    else:
        # ── Fallback path: no chunks found in primary collection ──────────────
        # Run a single Google Search-grounded Gemini call that populates ALL
        # required fields in one shot (no second supplementary call needed).
        print(f"    [INFO] No AlloyDB chunks for '{category}' patents of '{drug_name}' "
              f"— generating circumvention narrative via online search...")

        online_prompt = CIRCUMVENTION_ONLINE_PROMPT.format(
            claim_category=category,
            drug=drug_name,
            patent_numbers=", ".join(str(x) for x in patent_numbers),
        )
        try:
            online_resp = await gemini_client.aio.models.generate_content(
                model="gemini-2.5-flash",
                contents=online_prompt,
                config=search_config,  # Google Search grounding active
            )
            online_json = _extract_json_from_response(
                online_resp.text.strip() if online_resp.text else ""
            )
            if online_json:
                result = _repair_json(online_json)
                if isinstance(result, str):
                    result = json.loads(result)
                print(f"    [INFO] Online narrative generated successfully for '{category}'")
            else:
                raise ValueError("Empty or unparseable response from online search prompt")
        except Exception as e:
            print(f"    [WARN] Online circumvention search error: {e}")
            result = {
                "claim_category": category,
                "key_claim_limitations": [],
                "design_around_strategies": [],
                "white_space_opportunities": [],
                "overall_circumvention_difficulty": "N/A",
                "fda_precedents": [],
                "orange_book_gaps": [],
                "literature_alternatives": [],
                "regulatory_viability": f"Search error: {e}",
                "summary": (
                    "Patent text was unavailable in the local database and the "
                    "online search also failed. Manual review required."
                ),
            }

    result["patent_numbers"] = patent_numbers
    result["patent_count"]   = len(patents)
    print(f"    Found {len(result.get('design_around_strategies', []))} strategies")

    return category, result


async def run_circumvention_analysis(
    drug_name: str,
    patents_by_category: dict[str, list[dict]],
    chroma_client,
) -> dict:
    t0 = time.time()
    print(f"\n[Circumvention] Analysing design-around strategies for {drug_name}...")

    gemini_client = genai.Client(api_key=API_KEY)
    search_tools = [types.Tool(google_search=types.GoogleSearch())]
    search_config = types.GenerateContentConfig(tools=search_tools, temperature=0.1)

    semaphore = asyncio.Semaphore(CATEGORY_CONCURRENCY)

    async def _bounded_analyse(category, patents):
        async with semaphore:
            return await _analyse_one_category(
                drug_name, category, patents,
                chroma_client, gemini_client, search_config,
            )

    # Launch all categories concurrently (bounded by semaphore)
    tasks = [
        _bounded_analyse(category, patents)
        for category, patents in patents_by_category.items()
    ]
    results_list = await asyncio.gather(*tasks)

    category_results = {cat: res for cat, res in results_list}

    elapsed = round(time.time() - t0, 1)
    print(f"\n[Circumvention] Completed analysis for {drug_name} in {elapsed}s")

    return {
        "drug_name": drug_name,
        "categories_analysed": list(category_results.keys()),
        "results_by_category": category_results,
        "analysis_date": datetime.now().strftime("%Y-%m-%d"),
        "search_time_seconds": elapsed,
    }


def get_circumvention_for_drugs(non_blocking_df: pd.DataFrame, chroma_client,
                                max_patents_per_category: int = 10) -> dict[str, dict]:
    """Run circumvention analysis, limiting to max_patents_per_category patents per category."""
    cat_col = None
    for col_name in ["Step 1 Claim Category", "Patent Type"]:
        if col_name in non_blocking_df.columns:
            cat_col = col_name
            break

    if cat_col is None:
        print("[Circumvention] WARNING: No 'Step 1 Claim Category' column found.")
        return {}

    drugs_categories = {}
    for _, row in non_blocking_df.iterrows():
        drug       = str(row.get("Drug Name", "")).strip()
        pn         = str(row.get("Patent Number", "")).strip()
        cat        = str(row.get(cat_col, "")).strip()
        source_file = str(row.get("Source File", "")).strip()

        if not drug or not pn or cat in ("", "nan", "None"):
            continue
        if not source_file or source_file in ("nan", "None"):
            print(f"    [WARN] Missing 'Source File' for patent '{pn}' ({drug}) — skipping.")
            continue

        if drug not in drugs_categories:
            drugs_categories[drug] = {}
        if cat not in drugs_categories[drug]:
            drugs_categories[drug][cat] = []
        drugs_categories[drug][cat].append({
            "patent_number": pn,
            "drug_name": drug,
            "source_file": source_file,
        })

    # Limit patents per category
    for drug in drugs_categories:
        for cat in drugs_categories[drug]:
            patents = drugs_categories[drug][cat]
            if len(patents) > max_patents_per_category:
                print(f"    [{drug}/{cat}] Limiting from {len(patents)} to {max_patents_per_category} patents")
                drugs_categories[drug][cat] = patents[:max_patents_per_category]

    async def _run_all():
        # Run ALL drugs concurrently — each drug's categories are bounded
        # by CATEGORY_CONCURRENCY internally
        tasks = {
            drug: run_circumvention_analysis(drug, patents_by_cat, chroma_client)
            for drug, patents_by_cat in drugs_categories.items()
        }
        drug_names = list(tasks.keys())
        results_list = await asyncio.gather(*tasks.values())
        return dict(zip(drug_names, results_list))

    return asyncio.run(_run_all())


# ── BigQuery Writers ──────────────────────────────────────────────────────────

BQ_CIRC_TABLE   = "Circumvention_Table"
BQ_SCORE_TABLE  = "Patent_Thicket_Score_Table"


def _get_credentials():
    """Get credentials: use service account file if available, else default (Cloud Run)."""
    if CREDENTIALS_PATH and os.path.exists(CREDENTIALS_PATH):
        return service_account.Credentials.from_service_account_file(CREDENTIALS_PATH)
    return None  # Use ADC (Application Default Credentials)

def _get_bq_client() -> bigquery.Client:
    """Return an authenticated BigQuery client."""
    credentials = _get_credentials()
    return bigquery.Client(project=BQ_PROJECT_ID, credentials=credentials)


def write_circumvention_to_bq(circumvention_by_drug: dict):
    """
    Flattens circumvention analysis results into rows and writes them to
    BigQuery table: Circumvention_Table
    One row per design-around strategy (or one placeholder row if none exist).

    Columns
    -------
    Drug_Name, Patent_Category, Patents, Num_Patents,
    Overall_Difficulty, Strategy, Rationale, Feasibility,
    Regulatory_Pathway, Prior_Art_Support, Key_Claim_Limitations,
    White_Space_Opportunities, FDA_Precedents, Orange_Book_Gaps,
    Literature_Alternatives, Regulatory_Viability, Summary,
    Analysis_Date
    """
    rows = []
    for drug_name, circ_data in circumvention_by_drug.items():
        analysis_date = circ_data.get("analysis_date", "")
        for category, cat_result in circ_data.get("results_by_category", {}).items():
            strategies  = cat_result.get("design_around_strategies", [])
            difficulty  = cat_result.get("overall_circumvention_difficulty", "N/A")
            patent_nums = ", ".join(str(x) for x in cat_result.get("patent_numbers", []))
            patent_count = int(cat_result.get("patent_count", 0))
            limitations = "; ".join(str(x) for x in cat_result.get("key_claim_limitations", []))
            white_space = "; ".join(str(x) for x in cat_result.get("white_space_opportunities", []))
            fda_prec    = "; ".join(str(x) for x in cat_result.get("fda_precedents", []))
            ob_gaps     = "; ".join(str(x) for x in cat_result.get("orange_book_gaps", []))
            lit_alts    = "; ".join(str(x) for x in cat_result.get("literature_alternatives", []))
            reg_viab    = cat_result.get("regulatory_viability", "")
            summary     = cat_result.get("summary", "")

            common = dict(
                Drug_Name=drug_name,
                Patent_Category=category,
                Patents=patent_nums,
                Num_Patents=patent_count,
                Overall_Difficulty=difficulty,
                Key_Claim_Limitations=limitations,
                White_Space_Opportunities=white_space,
                FDA_Precedents=fda_prec,
                Orange_Book_Gaps=ob_gaps,
                Literature_Alternatives=lit_alts,
                Regulatory_Viability=reg_viab,
                Summary=summary,
                Analysis_Date=analysis_date,
            )

            if not strategies:
                rows.append({
                    **common,
                    "Strategy": "No strategies identified",
                    "Rationale": "",
                    "Feasibility": "",
                    "Regulatory_Pathway": "",
                    "Prior_Art_Support": "",
                })
            else:
                for strat in strategies:
                    rows.append({
                        **common,
                        "Strategy":           strat.get("strategy", ""),
                        "Rationale":          strat.get("rationale", ""),
                        "Feasibility":        strat.get("feasibility", ""),
                        "Regulatory_Pathway": strat.get("regulatory_pathway", ""),
                        "Prior_Art_Support":  strat.get("prior_art_support", ""),
                    })

    if not rows:
        print("[BQ] No circumvention rows to write.")
        return

    df_circ = pd.DataFrame(rows)
    table_ref = f"{BQ_PROJECT_ID}.{BQ_DATASET_ID}.{BQ_CIRC_TABLE}"
    client = _get_bq_client()

    # Align dtypes to match BQ schema
    try:
        table = client.get_table(table_ref)
        bq_type_map = {field.name: field.field_type for field in table.schema}
        for col in df_circ.columns:
            if col in bq_type_map:
                bq_type = bq_type_map[col]
                if bq_type in ("STRING",):
                    df_circ[col] = df_circ[col].astype(str).replace({"None": None, "nan": None, "NaN": None})
                elif bq_type in ("INTEGER", "INT64"):
                    df_circ[col] = pd.to_numeric(df_circ[col], errors="coerce").round(0).astype("Int64")
                elif bq_type in ("FLOAT", "FLOAT64", "NUMERIC"):
                    df_circ[col] = pd.to_numeric(df_circ[col], errors="coerce")
    except Exception as e:
        print(f"[BQ] Could not read schema for type alignment ({e}) — using autodetect")

    df_circ = df_circ.drop_duplicates()
    # created_at lets a re-run verify (via load_completed_drugs_from_bq below)
    # that a drug marked "completed" in the GCS checkpoint actually has rows
    # here — not just that the checkpoint claims it does.
    df_circ["created_at"] = pd.Timestamp.now(tz="UTC")

    # Skip rows where Drug_Name + Patent_Category already exist in BQ
    try:
        existing = client.query(
            f"SELECT DISTINCT Drug_Name, Patent_Category FROM `{table_ref}`"
        ).to_dataframe()
        if not existing.empty:
            existing["_key"] = (existing["Drug_Name"].astype(str).str.strip() + "|"
                                + existing["Patent_Category"].astype(str).str.strip())
            existing_keys = set(existing["_key"])
            df_circ["_key"] = (df_circ["Drug_Name"].astype(str).str.strip() + "|"
                               + df_circ["Patent_Category"].astype(str).str.strip())
            before = len(df_circ)
            df_circ = df_circ[~df_circ["_key"].isin(existing_keys)].drop(columns=["_key"])
            skipped = before - len(df_circ)
            if skipped > 0:
                print(f"[BQ] Skipped {skipped} circumvention rows (already in table)")
    except Exception as e:
        print(f"[BQ] Could not check existing circumvention rows ({e}) — writing all")

    if df_circ.empty:
        print(f"[BQ] All circumvention rows already exist — nothing to write.")
        return

    job_config = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
        autodetect=True,
        schema_update_options=[bigquery.SchemaUpdateOption.ALLOW_FIELD_ADDITION],
    )
    job = client.load_table_from_dataframe(df_circ, table_ref, job_config=job_config,
                                           location=BQ_LOCATION)
    job.result()
    print(f"[BQ] Circumvention_Table: {len(df_circ)} new rows written → {table_ref}")


def write_score_to_bq(drug_scores: list, refresh=False):
    """
    Flattens patent thicket score data into rows and writes them to
    BigQuery table: Patent_Thicket_Score_Table
    One row per drug+jurisdiction, plus one 'Final Score (Average)' row per drug.

    Columns
    -------
    Drug_Name, Jurisdiction, Combined_Total, Adjusted_Count,
    Active_Technology_Areas, Active_Categories,
    Density_Interpretation, Diversity_Interpretation,
    Density_Score, Diversity_Score, Base_Score,
    Validation_Pct, Final_Score, Score_Label
    """
    rows = []
    for sd in drug_scores:
        for jd in sd.get("jurisdiction_scores", []):
            rows.append({
                "Drug_Name":               sd["drug"],
                "Jurisdiction":            jd["jurisdiction"],
                "Combined_Total":          jd["combined_total"],
                "Adjusted_Count":          jd["adjusted_count"],
                "Active_Technology_Areas": jd["active_areas"],
                "Active_Categories":       jd["active_categories"],
                "Density_Interpretation":  jd["density_label"],
                "Diversity_Interpretation": jd["diversity_label"],
                "Density_Score":           jd["density_score"],
                "Diversity_Score":         jd["diversity_score"],
                "Base_Score":              jd["base_score"],
                "Validation_Pct":          jd["validation_pct"],
                "Final_Score":             jd["final_score"],
                "Score_Label":             jd["final_label"],
            })

        # Average row across all jurisdictions
        avg_final     = sd["avg_final_score"]
        avg_final_int = max(1, min(5, round(avg_final)))
        rows.append({
            "Drug_Name":               sd["drug"],
            "Jurisdiction":            "Final Score (Average)",
            "Combined_Total":          None,
            "Adjusted_Count":          None,
            "Active_Technology_Areas": None,
            "Active_Categories":       "",
            "Density_Interpretation":  "",
            "Diversity_Interpretation": "",
            "Density_Score":           None,
            "Diversity_Score":         None,
            "Base_Score":              None,
            "Validation_Pct":          None,
            "Final_Score":             avg_final,
            "Score_Label":             THICKET_SCORE_LABELS.get(avg_final_int, ""),
        })

    if not rows:
        print("[BQ] No score rows to write.")
        return

    df_score = pd.DataFrame(rows)
    table_ref = f"{BQ_PROJECT_ID}.{BQ_DATASET_ID}.{BQ_SCORE_TABLE}"
    client = _get_bq_client()

    # Align dtypes to match BQ schema — read the table schema and cast accordingly
    try:
        table = client.get_table(table_ref)
        bq_type_map = {field.name: field.field_type for field in table.schema}
        for col in df_score.columns:
            if col in bq_type_map:
                bq_type = bq_type_map[col]
                if bq_type in ("STRING",):
                    df_score[col] = df_score[col].astype(str).replace({"None": None, "nan": None, "NaN": None})
                elif bq_type in ("INTEGER", "INT64"):
                    df_score[col] = pd.to_numeric(df_score[col], errors="coerce").round(0).astype("Int64")
                elif bq_type in ("FLOAT", "FLOAT64", "NUMERIC"):
                    df_score[col] = pd.to_numeric(df_score[col], errors="coerce")
    except Exception as e:
        print(f"[BQ] Could not read schema for type alignment ({e}) — using autodetect")

    df_score = df_score.drop_duplicates()
    df_score["created_at"] = pd.Timestamp.now(tz="UTC")

    if refresh:
        # Delete existing score rows for the drugs we're about to write
        drug_names = df_score["Drug_Name"].dropna().unique().tolist()
        for dn in drug_names:
            try:
                delete_sql = f"DELETE FROM `{table_ref}` WHERE Drug_Name = @drug"
                job_config = bigquery.QueryJobConfig(
                    query_parameters=[bigquery.ScalarQueryParameter("drug", "STRING", dn)]
                )
                client.query(delete_sql, job_config=job_config).result()
                print(f"[REFRESH] Deleted existing score rows for '{dn}'")
            except Exception as e:
                print(f"[REFRESH] Could not delete rows for '{dn}': {e}")
    else:
        # Skip rows where Drug_Name + Jurisdiction already exist in BQ
        try:
            existing = client.query(
                f"SELECT DISTINCT Drug_Name, Jurisdiction FROM `{table_ref}`"
            ).to_dataframe()
            if not existing.empty:
                existing["_key"] = (existing["Drug_Name"].astype(str).str.strip() + "|"
                                    + existing["Jurisdiction"].astype(str).str.strip())
                existing_keys = set(existing["_key"])
                df_score["_key"] = (df_score["Drug_Name"].astype(str).str.strip() + "|"
                                    + df_score["Jurisdiction"].astype(str).str.strip())
                before = len(df_score)
                df_score = df_score[~df_score["_key"].isin(existing_keys)].drop(columns=["_key"])
                skipped = before - len(df_score)
                if skipped > 0:
                    print(f"[BQ] Skipped {skipped} score rows (already in table)")
        except Exception as e:
            print(f"[BQ] Could not check existing score rows ({e}) — writing all")

    if df_score.empty:
        print(f"[BQ] All score rows already exist — nothing to write.")
        return

    job_config = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
        autodetect=True,
        schema_update_options=[bigquery.SchemaUpdateOption.ALLOW_FIELD_ADDITION],
    )
    job = client.load_table_from_dataframe(df_score, table_ref, job_config=job_config,
                                           location=BQ_LOCATION)
    job.result()
    print(f"[BQ] Patent_Thicket_Score_Table: {len(df_score)} new rows written → {table_ref}")


# ── Main ──────────────────────────────────────────────────────────────────────
EXCLUDED_CATEGORIES = {"Composition Of Matter"}

def process_patents(skip_circumvention=False, drug_filter=None, refresh_scores=False,
                    rerun=False, max_patents_per_category=10):
    # ── Load from BigQuery ────────────────────────────────────────────────
    df = load_data_from_bigquery()
    df.columns = df.columns.str.strip()
    df = df.drop_duplicates()

    # All values arrive as STR from BQ; coerce required text columns
    for col in ["Tag", "Step 1 Claim Category", "Phase", "Drug Name"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
        else:
            print(f"  [WARN] Expected column '{col}' not found in BigQuery table.")

    # Normalize category casing to title-case so variants like
    # "Method of Treatment" and "Method of treatment" collapse to one.
    df["Step 1 Claim Category"] = df["Step 1 Claim Category"].str.title()

    # Normalize Jurisdiction column if present
    if "Jurisdiction" in df.columns:
        df["Jurisdiction"] = df["Jurisdiction"].astype(str).str.strip()

    # Normalize Type column if present
    if "Type" in df.columns:
        df["Type"] = df["Type"].astype(str).str.strip()

    # Ensure "No Of Forecasted Patents" column is numeric (default 0)
    # Try to find the column with flexible matching (case/whitespace variations)
    forecasted_col = None
    for c in df.columns:
        if c.lower().replace(" ", "").replace("_", "") in (
            "noofforecastedpatents", "numberofforecastedpatents",
            "forecastedpatents", "noofforecasted",
        ):
            forecasted_col = c
            break
    if forecasted_col is None and "No Of Forecasted Patents" in df.columns:
        forecasted_col = "No Of Forecasted Patents"

    if forecasted_col is not None:
        print(f"  [INFO] Found forecasted column: '{forecasted_col}'")
        df["No Of Forecasted Patents"] = pd.to_numeric(
            df[forecasted_col], errors="coerce"
        ).fillna(0).astype(int)
        print(f"  [INFO] Forecasted patent values (unique): {sorted(df['No Of Forecasted Patents'].unique())}")
    else:
        print(f"  [WARN] 'No Of Forecasted Patents' column not found. Available columns: {list(df.columns)}")
        df["No Of Forecasted Patents"] = 0

    # Check Type column values
    if "Type" in df.columns:
        print(f"  [INFO] Type column values (unique): {df['Type'].unique().tolist()}")

    all_categories = sorted(
        c for c in df["Step 1 Claim Category"].replace("nan", pd.NA).dropna().unique()
        if c not in EXCLUDED_CATEGORIES
    )

    non_blocking = df[df["Tag"].str.upper() == "NON-BLOCKING"].copy()

    if drug_filter:
        non_blocking = non_blocking[
            non_blocking["Drug Name"].str.lower() == drug_filter.lower()
        ].copy()
        if non_blocking.empty:
            print(f"No NON-BLOCKING rows found for drug: '{drug_filter}'")
            return
        print(f"[Filter] Processing single drug: '{drug_filter}' "
              f"({len(non_blocking)} NON-BLOCKING rows)")

    if non_blocking.empty:
        print("No rows found with Tag = 'NON-BLOCKING'.")
        return

    # ── GLP-1 drug-universe guard ───────────────────────────────────────────
    if drug_filter:
        if not glp1_universe.require_allowed_drug(drug_filter):
            return
    else:
        before_names = sorted(non_blocking["Drug Name"].dropna().unique())
        allowed_names = glp1_universe.filter_allowed_drugs(before_names)
        if len(allowed_names) != len(before_names):
            allowed_norm = {glp1_universe.normalize(d) for d in allowed_names}
            non_blocking = non_blocking[
                non_blocking["Drug Name"].apply(lambda d: glp1_universe.normalize(d) in allowed_norm)
            ].copy()
            if non_blocking.empty:
                print("No GLP-1 drugs remain after applying the drug-universe filter.")
                return

    drugs = non_blocking["Drug Name"].unique()

    # ── GCS Checkpoint: skip drugs already completed ──────────────────────
    _ipd3_ckpt_subfolder = "ipd3_checkpoints"
    _ipd3_ckpt_file = "completed_drugs.json"
    completed_drugs = set()

    # Allow clearing the checkpoint via env var (useful after a crash that
    # left the checkpoint in an inconsistent state, e.g. score data computed
    # but never written to BQ). Set CLEAR_IPD3_CHECKPOINT=true on Cloud Run.
    _force_clear = os.getenv("CLEAR_IPD3_CHECKPOINT", "").lower() in ("1", "true", "yes")

    if refresh_scores or rerun or _force_clear:
        reason = ("CLEAR_IPD3_CHECKPOINT env var" if _force_clear
                  else "RERUN" if rerun else "REFRESH")
        print(f"[{reason}] Bypassing checkpoint, will reprocess all drugs")
    else:
        try:
            from cog import gcs_cache
            ckpt_data = gcs_cache.read_json(_ipd3_ckpt_subfolder, _ipd3_ckpt_file)
            if ckpt_data and isinstance(ckpt_data, list):
                completed_drugs = set(ckpt_data)
                print(f"[CHECKPOINT] {len(completed_drugs)} drug(s) already done in GCS: {completed_drugs}")
        except Exception as e:
            print(f"[CHECKPOINT] Could not load ipd3 checkpoint: {e}")

    # ── Cross-check against BigQuery: a drug is only truly "done" if BOTH
    # its circumvention rows AND score rows actually made it into their
    # respective BQ tables. The checkpoint is saved after the in-memory
    # analysis — BEFORE the BQ writes — so if either write crashed (e.g.
    # the UnboundLocalError in write_score_to_bq), the drug would otherwise
    # be skipped forever on every future run. We verify against both tables.
    if completed_drugs and not skip_circumvention:
        try:
            client = _get_bq_client()
            circ_table_ref  = f"{BQ_PROJECT_ID}.{BQ_DATASET_ID}.{BQ_CIRC_TABLE}"
            score_table_ref = f"{BQ_PROJECT_ID}.{BQ_DATASET_ID}.{BQ_SCORE_TABLE}"

            existing_circ = client.query(
                f"SELECT DISTINCT Drug_Name FROM `{circ_table_ref}`"
            ).to_dataframe()
            drugs_with_circ = set(existing_circ["Drug_Name"].astype(str).str.strip()) if not existing_circ.empty else set()

            existing_score = client.query(
                f"SELECT DISTINCT Drug_Name FROM `{score_table_ref}`"
            ).to_dataframe()
            drugs_with_score = set(existing_score["Drug_Name"].astype(str).str.strip()) if not existing_score.empty else set()

            # A drug is truly done only if it has rows in BOTH tables
            drugs_fully_done = drugs_with_circ & drugs_with_score
            missing = completed_drugs - drugs_fully_done
            if missing:
                print(f"[CHECKPOINT] {len(missing)} drug(s) marked done but missing from "
                      f"{BQ_CIRC_TABLE} and/or {BQ_SCORE_TABLE} — will reprocess: {sorted(missing)}")
                completed_drugs -= missing
        except Exception as e:
            print(f"[CHECKPOINT] Could not verify against BQ tables ({e}) "
                  f"— trusting the GCS checkpoint as-is")

    def _save_ipd3_checkpoint():
        try:
            from cog import gcs_cache
            gcs_cache.write_json(_ipd3_ckpt_subfolder, _ipd3_ckpt_file, sorted(completed_drugs))
        except Exception as e:
            print(f"[CHECKPOINT] GCS write failed: {e}")

    drug_scores = []

    for drug in drugs:
        if drug in completed_drugs:
            print(f"\n  [SKIP] {drug} — already completed (checkpoint)")
            continue

        drug_df = non_blocking[non_blocking["Drug Name"] == drug].copy()

        # ── Build forecasted source (BLOCKING rows with Type == "Forecasted") ──
        forecasted_drug_df = df[
            (df["Drug Name"] == drug) &
            (df["Type"].str.lower() == "forecasted")
        ].copy()
        print(f"  [INFO] Drug '{drug}': {len(forecasted_drug_df)} forecasted rows found")

        filtered_cats = [c for c in all_categories if c not in EXCLUDED_CATEGORIES]

        def compute_block_scores(block_df, forecasted_source_df):
            """Compute score info dict for a given df block (jurisdiction or all)."""
            category_counts = block_df.groupby("Step 1 Claim Category").size().to_dict()
            category_counts = {cat: category_counts.get(cat, 0) for cat in filtered_cats}

            if not forecasted_source_df.empty:
                forecasted_counts = (
                    forecasted_source_df
                    .groupby("Step 1 Claim Category")["No Of Forecasted Patents"]
                    .max()
                    .to_dict()
                )
            else:
                forecasted_counts = {}
            forecasted_counts = {cat: int(forecasted_counts.get(cat, 0)) for cat in filtered_cats}

            total_patents    = sum(category_counts.values())
            forecasted_total = sum(forecasted_counts.values())
            combined_total   = total_patents + forecasted_total
            adjusted_count   = total_patents + (forecasted_total * 0.5)

            active_cat_names = [cat for cat in filtered_cats if category_counts[cat] > 0]
            active_areas     = len(active_cat_names)

            score_info = compute_final_score(
                total_patents, combined_total, adjusted_count, active_areas
            )
            return (score_info, active_cat_names, combined_total, adjusted_count, active_areas)

        # ── Per-Jurisdiction score computation (skip WO) ──────────────────────
        jurisdiction_scores = []

        if "Jurisdiction" in drug_df.columns:
            jurisdictions = sorted(
                j for j in drug_df["Jurisdiction"].dropna().unique()
                if j not in ("", "nan", "None") and j.upper() != "WO"
            )
            for jurisdiction in jurisdictions:
                jur_df = drug_df[drug_df["Jurisdiction"] == jurisdiction].copy()
                if "Jurisdiction" in forecasted_drug_df.columns:
                    jur_forecasted_df = forecasted_drug_df[
                        forecasted_drug_df["Jurisdiction"] == jurisdiction
                    ].copy()
                else:
                    jur_forecasted_df = forecasted_drug_df.copy()

                j_score_info, j_cats, j_combined, j_adjusted, j_active_areas = \
                    compute_block_scores(jur_df, jur_forecasted_df)

                j_label1 = patent_count_label(j_score_info["effective_adjusted_count"])
                j_label2 = diversity_label(j_active_areas)

                jurisdiction_scores.append({
                    "drug":             drug,
                    "jurisdiction":     jurisdiction,
                    "combined_total":   j_combined,
                    "adjusted_count":   j_adjusted,
                    "active_areas":     j_active_areas,
                    "active_categories": ", ".join(j_cats) if j_cats else "None",
                    "density_label":    j_label1,
                    "diversity_label":  j_label2,
                    "density_score":    j_score_info["density_score"],
                    "diversity_score":  j_score_info["diversity_score"],
                    "base_score":       j_score_info["base_score"],
                    "validation_pct":   j_score_info["validation_pct"],
                    "final_score":      j_score_info["final_score"],
                    "final_label":      j_score_info["final_label"],
                })

        # ── Average final score across jurisdictions ───────────────────────────
        if jurisdiction_scores:
            avg_final = round(
                sum(js["final_score"] for js in jurisdiction_scores)
                / len(jurisdiction_scores), 1
            )
        else:
            avg_final = 0

        # ── Overall counts (non-WO, for drug_scores summary) ──────────────────
        non_wo_df = drug_df[drug_df["Jurisdiction"].str.upper() != "WO"] \
            if "Jurisdiction" in drug_df.columns else drug_df
        _counts = non_wo_df.groupby("Step 1 Claim Category").size().to_dict()
        total_patents = sum(_counts.get(c, 0) for c in all_categories if c not in EXCLUDED_CATEGORIES)
        active_areas  = sum(1 for c in all_categories
                            if c not in EXCLUDED_CATEGORIES and _counts.get(c, 0) > 0)
        active_cat_names = [c for c in all_categories
                            if c not in EXCLUDED_CATEGORIES and _counts.get(c, 0) > 0]

        drug_scores.append({
            "drug":               drug,
            "total_patents":      total_patents,
            "active_areas":       active_areas,
            "active_categories":  ", ".join(active_cat_names) if active_cat_names else "None",
            "combined_total":     sum(js["combined_total"] for js in jurisdiction_scores),
            "adjusted_count":     sum(js["adjusted_count"] for js in jurisdiction_scores),
            "avg_final_score":    avg_final,
            "jurisdiction_scores": jurisdiction_scores,
        })

        # Mark drug as completed in checkpoint
        completed_drugs.add(drug)
        _save_ipd3_checkpoint()

    # ── Circumvention Analysis → BigQuery ────────────────────────────────
    # Only run for drugs that were actually processed (not skipped by checkpoint)
    processed_drugs = {ds["drug"] for ds in drug_scores}
    if not skip_circumvention and processed_drugs:
        if not API_KEY:
            print("\n⚠️  GEMINI_API_KEY not set — skipping circumvention analysis.")
        else:
            print(f"\n{'='*60}")
            print(f"Running circumvention / 505(b)(2) design-around analysis...")
            print(f"{'='*60}")
            # Filter non_blocking to only the drugs we actually scored
            nb_to_analyse = non_blocking[non_blocking["Drug Name"].isin(processed_drugs)]
            chroma_client = get_chroma_clients()[0]
            circumvention_by_drug = get_circumvention_for_drugs(
                nb_to_analyse, chroma_client,
                max_patents_per_category=max_patents_per_category,
            )
            if circumvention_by_drug:
                write_circumvention_to_bq(circumvention_by_drug)
    elif skip_circumvention:
        print("\n⏭️  Circumvention analysis skipped (--skip-circumvention flag).")
    else:
        print("\n⏭️  No new drugs to analyse — circumvention skipped.")

    # ── Score → BigQuery ──────────────────────────────────────────────────
    write_score_to_bq(drug_scores, refresh=refresh_scores)

    print(f"\nDone! Results written to BigQuery dataset: {BQ_DATASET_ID}")
    print(f"  Tables: {BQ_CIRC_TABLE}, {BQ_SCORE_TABLE}")
    print(f"  {len(non_blocking)} NON-BLOCKING rows | {len(drugs)} drug(s): {', '.join(str(d) for d in drugs)}")
    print(f"  {len(all_categories)} Step 1 Claim Categories: {', '.join(str(c) for c in all_categories)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Patent Thicket Analysis & Circumvention Strategy (505(b)(2))"
    )
    parser.add_argument(
        "drug", nargs="?", default=None,
        help="Optional: single drug name to process (e.g. Semaglutide). Omit to process all drugs.",
    )
    parser.add_argument(
        "--skip-circumvention", action="store_true",
        help="Skip circumvention / 505(b)(2) design-around analysis",
    )
    parser.add_argument(
        "--refresh-scores", action="store_true",
        help="Delete existing score rows for the drug(s) and recompute. Bypasses checkpoint.",
    )
    parser.add_argument(
        "--rerun", action="store_true",
        help="Bypass checkpoint and rerun everything (scoring + circumvention).",
    )
    parser.add_argument(
        "--max-patents-per-category", type=int, default=10,
        help="Max patents per category for circumvention analysis (default: 10).",
    )
    args = parser.parse_args()

    process_patents(
        args.skip_circumvention,
        drug_filter=args.drug,
        refresh_scores=args.refresh_scores or args.rerun,
        rerun=args.rerun,
        max_patents_per_category=args.max_patents_per_category,
    )
