"""
IPD3 Orchestrator — Run Gemini & Claude Pipelines, Then LLM-as-Judge Evaluation

This orchestrator:
  1. Runs ipd3bq__6_.py (Gemini) — writes to BQ + local JSON
  2. Runs ipd3bq_claude.py (Claude) — writes to BQ + local JSON
  3. Loads both results from LOCAL JSON files (no BQ queries needed)
  4. Compares circumvention strategies & thicket scores using Claude Sonnet 4.6 as judge
  5. Writes evaluation to Excel (default) or BQ (with --to-bq)

No BigQuery jobs are created for evaluation. Pipelines write JSON locally
and the orchestrator reads from those files.

Usage:
    python ipd3bq_orchestrator.py                              # run all, eval → Excel
    python ipd3bq_orchestrator.py --drug Semaglutide           # single drug
    python ipd3bq_orchestrator.py --skip-run                   # skip pipelines, eval only
    python ipd3bq_orchestrator.py --skip-gemini                # skip Gemini, run Claude + eval
    python ipd3bq_orchestrator.py --skip-claude                # skip Claude, run Gemini + eval
    python ipd3bq_orchestrator.py --to-bq                      # write eval results to BQ too
    python ipd3bq_orchestrator.py --no-bq                      # pass --no-bq to pipelines too
"""

import os
import re
import sys
import json
import time
import random
import argparse
import subprocess
from datetime import datetime
from typing import Any, Dict, List, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
from anthropic import AnthropicVertex, RateLimitError

try:
    from dotenv import load_dotenv
    load_dotenv(override=True)
except ImportError:
    pass

try:
    from json_repair import repair_json as _repair_json_lib
except ImportError:
    _repair_json_lib = None

# ── Config ─────────────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

BQ_PROJECT_ID = os.getenv("BQ_PROJECT_ID", "cognito-prod-394707")
BQ_DATASET_ID = os.getenv("BQ_DATASET_ID", "cognito_prod_datamart")
BQ_LOCATION   = os.getenv("BQ_LOCATION", "asia-south1")
GCP_REGION    = os.getenv("GCP_REGION", "us-east5")
PROJECT_ID    = os.getenv("BQ_PROJECT_ID", "cognito-prod-394707")

CREDENTIALS_PATH = os.environ.get("CREDENTIALS_PATH") or os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "")
if CREDENTIALS_PATH and not os.path.isabs(CREDENTIALS_PATH):
    CREDENTIALS_PATH = os.path.join(SCRIPT_DIR, CREDENTIALS_PATH)

GCS_BUCKET    = os.getenv("GCS_BUCKET", "cognito-gcs")
GCS_BASE_PATH = os.getenv("GCS_EVAL_PATH", "Cognito_new/eval_reports")

CLAUDE_MODEL   = os.getenv("CLAUDE_MODEL", "claude-sonnet-4-6")
MAX_WORKERS    = 6
LLM_MAX_TOKENS = 16384

OUTPUT_DIR = os.getenv("IPD3_OUTPUT_DIR", os.path.join(SCRIPT_DIR, "ipd3_output"))

EVAL_TABLE = "IPD3_Eval_Table"


# ── Claude Client ─────────────────────────────────────────────────────────────

def get_claude_client() -> AnthropicVertex:
    return AnthropicVertex(region=GCP_REGION, project_id=PROJECT_ID)


def _call_claude(client: AnthropicVertex, prompt: str, max_retries: int = 3) -> Dict[str, Any]:
    for attempt in range(max_retries + 1):
        try:
            resp = client.messages.create(
                model=CLAUDE_MODEL, max_tokens=LLM_MAX_TOKENS,
                messages=[{"role": "user", "content": prompt}],
            )
            text = resp.content[0].text.strip()
            text = re.sub(r"^```(?:json)?", "", text).strip()
            text = re.sub(r"```$", "", text).strip()
            try:
                return json.loads(text)
            except json.JSONDecodeError:
                if _repair_json_lib:
                    return json.loads(_repair_json_lib(text))
                text = re.sub(r",\s*}", "}", text)
                text = re.sub(r",\s*]", "]", text)
                return json.loads(text)
        except RateLimitError:
            if attempt == max_retries: raise
            wait = 2.0 * (2 ** attempt) * (1 + random.uniform(0, 0.25))
            print(f"    [Claude] Rate limited, retry {attempt+1}/{max_retries} in {wait:.1f}s")
            time.sleep(wait)
        except Exception as e:
            if attempt == max_retries:
                return {"error": str(e)}
            time.sleep(2.0 * (2 ** attempt))
    return {}


# ── Step 1 & 2: Run Pipelines ────────────────────────────────────────────────

def _find_script(name):
    """Find a script by name in SCRIPT_DIR."""
    candidates = [name, name.replace("(", "_").replace(")", "_")]
    for c in candidates:
        p = os.path.join(SCRIPT_DIR, c)
        if os.path.exists(p):
            return p
    return os.path.join(SCRIPT_DIR, name)


def _build_pipeline_cmd(script_name, drug=None, extra_args=None):
    script = _find_script(script_name)
    cmd = [sys.executable, script]
    if drug: cmd.append(drug)
    if extra_args: cmd.extend(extra_args)
    return cmd


def run_pipelines_parallel(drug=None, extra_args=None,
                           skip_gemini=False, skip_claude=False):
    """Run Gemini and Claude pipelines in parallel using subprocess.Popen."""
    procs = {}

    if not skip_gemini:
        cmd_g = _build_pipeline_cmd("ipd3bq__6_.py", drug, extra_args)
        print(f"\n{'='*60}")
        print(f"STEP 1a: Launching Gemini pipeline (parallel)")
        print(f"  Command: {' '.join(cmd_g)}")
        print(f"{'='*60}")
        procs["Gemini"] = subprocess.Popen(cmd_g)
    else:
        print("\n⏭️  Gemini pipeline skipped")

    if not skip_claude:
        cmd_c = _build_pipeline_cmd("ipd3bq_claude.py", drug, extra_args)
        print(f"\n{'='*60}")
        print(f"STEP 1b: Launching Claude pipeline (parallel)")
        print(f"  Command: {' '.join(cmd_c)}")
        print(f"{'='*60}")
        procs["Claude"] = subprocess.Popen(cmd_c)
    else:
        print("\n⏭️  Claude pipeline skipped")

    # Wait for both to finish
    if procs:
        print(f"\n  Waiting for {len(procs)} pipeline(s) to complete ...")
    for name, proc in procs.items():
        rc = proc.wait()
        if rc != 0:
            print(f"[WARN] {name} pipeline exited with code {rc}")
        else:
            print(f"  [{name}] Done (exit 0)")


# ── Step 3: Load Results from LOCAL JSON ──────────────────────────────────────

def _load_json_safe(path: str) -> List[Dict]:
    """Load a JSON file or return empty list if missing/corrupt."""
    if not os.path.exists(path):
        print(f"[WARN] File not found: {path}")
        return []
    try:
        with open(path) as f:
            data = json.load(f)
        print(f"[JSON] Loaded {len(data)} rows from {path}")
        return data
    except Exception as e:
        print(f"[WARN] Failed to load {path}: {e}")
        return []


def load_circumvention_results(drug=None) -> pd.DataFrame:
    """Load Gemini + Claude circumvention results from local JSON, joined by Drug+Category."""
    gemini_rows = _load_json_safe(os.path.join(OUTPUT_DIR, "circumvention_gemini.json"))
    claude_rows = _load_json_safe(os.path.join(OUTPUT_DIR, "circumvention_claude.json"))

    if not gemini_rows and not claude_rows:
        return pd.DataFrame()

    g_df = pd.DataFrame(gemini_rows) if gemini_rows else pd.DataFrame()
    c_df = pd.DataFrame(claude_rows) if claude_rows else pd.DataFrame()

    # Prefix columns
    if not g_df.empty:
        g_key = g_df[["Drug_Name", "Patent_Category"]].copy()
        g_renamed = g_df.rename(columns={c: f"gemini_{c}" for c in g_df.columns if c not in ("Drug_Name", "Patent_Category")})
    else:
        g_renamed = pd.DataFrame()

    if not c_df.empty:
        c_renamed = c_df.rename(columns={c: f"claude_{c}" for c in c_df.columns if c not in ("Drug_Name", "Patent_Category")})
    else:
        c_renamed = pd.DataFrame()

    # Join
    if not g_renamed.empty and not c_renamed.empty:
        merged = pd.merge(g_renamed, c_renamed, on=["Drug_Name", "Patent_Category"], how="outer")
    elif not g_renamed.empty:
        merged = g_renamed
    else:
        merged = c_renamed

    if drug and not merged.empty:
        merged = merged[merged["Drug_Name"].str.lower() == drug.lower()]

    print(f"[LOAD] {len(merged)} circumvention comparison rows")
    return merged



# ── LLM-as-Judge Prompts ─────────────────────────────────────────────────────

CIRCUMVENTION_EVAL_PROMPT = """
You are a senior pharmaceutical patent attorney evaluating the quality of an LLM's
circumvention analysis against a GROUND TRUTH reference.

The GROUND TRUTH was produced by Claude Sonnet 4.6. You are evaluating whether
Gemini 2.5 Flash's output is correct and complete when measured against this reference.

Drug: {drug}
Patent Category: {patent_category}
Patents: {patents}

--- GROUND TRUTH (Claude Sonnet 4.6) ---
Difficulty: {claude_difficulty}
Strategy: {claude_strategy}
Rationale: {claude_rationale}
Feasibility: {claude_feasibility}
Regulatory Pathway: {claude_regulatory_pathway}
Prior Art Support: {claude_prior_art_support}
Key Claim Limitations: {claude_key_claim_limitations}
White Space: {claude_white_space}
FDA Precedents: {claude_fda_precedents}
Summary: {claude_summary}
--- END GROUND TRUTH ---

--- GEMINI OUTPUT (under evaluation) ---
Difficulty: {gemini_difficulty}
Strategy: {gemini_strategy}
Rationale: {gemini_rationale}
Feasibility: {gemini_feasibility}
Regulatory Pathway: {gemini_regulatory_pathway}
Prior Art Support: {gemini_prior_art_support}
Key Claim Limitations: {gemini_key_claim_limitations}
White Space: {gemini_white_space}
FDA Precedents: {gemini_fda_precedents}
Summary: {gemini_summary}
--- END GEMINI ---

Compare Gemini's output against the Claude ground truth. Score Gemini on:
1. **Faithfulness**: Does Gemini avoid hallucinated references, fabricated patent details,
   or unsupported assertions not present in the ground truth?
2. **Grounding**: Can Gemini's claims be traced to the patent data? Does it cite the same
   or equivalent source material as the ground truth?
3. **Relevance**: Is Gemini's analysis specific to this drug/category, or generic boilerplate?
4. **Accuracy**: Are Gemini's claim limitations and strategies technically/legally correct
   when compared to the ground truth?
5. **Completeness**: Does Gemini cover all circumvention approaches identified in the ground truth?
6. **Feasibility**: Are Gemini's feasibility ratings consistent with the ground truth?
7. **Regulatory Viability**: Are Gemini's 505(b)(2) pathways realistic per the ground truth?

Respond ONLY with valid JSON:
{{
  "agreement": <true if Gemini's output substantially agrees with ground truth, false otherwise>,
  "faithfulness_score": <integer 1-5, Gemini's score>,
  "grounding_score": <integer 1-5, Gemini's score>,
  "relevance_score": <integer 1-5, Gemini's score>,
  "accuracy_score": <integer 1-5, Gemini's score>,
  "completeness_score": <integer 1-5, Gemini's score>,
  "feasibility_score": <integer 1-5, Gemini's score>,
  "regulatory_score": <integer 1-5, Gemini's score>,
  "faithfulness_notes": "<1-2 sentences on Gemini's hallucination level vs ground truth>",
  "grounding_notes": "<1-2 sentences on whether Gemini's claims trace to source data>",
  "relevance_notes": "<1-2 sentences on Gemini's specificity vs ground truth>",
  "deviation_explanation": "<2-3 sentences: where and how Gemini deviates from the ground truth, or 'No significant deviations' if agreement is true>",
  "overall_assessment": "<2-3 sentence assessment of Gemini's output quality against the ground truth>"
}}
"""

OVERALL_SYNTHESIS_PROMPT = """
You are a senior pharmaceutical patent attorney providing a final synthesis.

Claude Sonnet 4.6's output is the GROUND TRUTH. You evaluated Gemini 2.5 Flash against it.

Drug: {drug}

Summary statistics:
- Categories evaluated: {num_categories}
- Gemini agreed with ground truth: {agreed} / {num_categories}
- Avg Faithfulness: {avg_faithfulness}
- Avg Grounding: {avg_grounding}
- Avg Relevance: {avg_relevance}
- Avg Accuracy: {avg_accuracy}
- Avg Completeness: {avg_completeness}

Respond ONLY with valid JSON:
{{
  "gemini_overall_correct": <true if Gemini mostly agrees with ground truth, false otherwise>,
  "confidence": "<high|medium|low>",
  "key_strengths": ["<where Gemini matched or exceeded ground truth>"],
  "key_weaknesses": ["<where Gemini deviated from ground truth>"],
  "recommendation": "<2-3 sentence verdict on Gemini's reliability for this drug>"
}}
"""


# ── Evaluation Logic ──────────────────────────────────────────────────────────

def _safe_get(row, key, default="N/A"):
    v = row.get(key)
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return default
    return str(v)[:500]


def evaluate_circumvention_row(client, row):
    prompt = CIRCUMVENTION_EVAL_PROMPT.format(
        drug=_safe_get(row, "Drug_Name"),
        patent_category=_safe_get(row, "Patent_Category"),
        patents=_safe_get(row, "gemini_Patents") or _safe_get(row, "claude_Patents"),
        gemini_difficulty=_safe_get(row, "gemini_Overall_Difficulty"),
        gemini_strategy=_safe_get(row, "gemini_Strategy"),
        gemini_rationale=_safe_get(row, "gemini_Rationale"),
        gemini_feasibility=_safe_get(row, "gemini_Feasibility"),
        gemini_regulatory_pathway=_safe_get(row, "gemini_Regulatory_Pathway"),
        gemini_prior_art_support=_safe_get(row, "gemini_Prior_Art_Support"),
        gemini_key_claim_limitations=_safe_get(row, "gemini_Key_Claim_Limitations"),
        gemini_white_space=_safe_get(row, "gemini_White_Space_Opportunities"),
        gemini_fda_precedents=_safe_get(row, "gemini_FDA_Precedents"),
        gemini_summary=_safe_get(row, "gemini_Summary"),
        claude_difficulty=_safe_get(row, "claude_Overall_Difficulty"),
        claude_strategy=_safe_get(row, "claude_Strategy"),
        claude_rationale=_safe_get(row, "claude_Rationale"),
        claude_feasibility=_safe_get(row, "claude_Feasibility"),
        claude_regulatory_pathway=_safe_get(row, "claude_Regulatory_Pathway"),
        claude_prior_art_support=_safe_get(row, "claude_Prior_Art_Support"),
        claude_key_claim_limitations=_safe_get(row, "claude_Key_Claim_Limitations"),
        claude_white_space=_safe_get(row, "claude_White_Space_Opportunities"),
        claude_fda_precedents=_safe_get(row, "claude_FDA_Precedents"),
        claude_summary=_safe_get(row, "claude_Summary"),
    )
    return _call_claude(client, prompt)


def run_overall_synthesis(client, drug, circ_evals):
    def _avg(evals, key):
        vals = [e.get(key, 0) for e in evals if e.get(key)]
        return round(sum(vals) / len(vals), 1) if vals else 0

    agreed = sum(1 for e in circ_evals if e.get("agreement"))

    prompt = OVERALL_SYNTHESIS_PROMPT.format(
        drug=drug, num_categories=len(circ_evals), agreed=agreed,
        avg_faithfulness=_avg(circ_evals, "faithfulness_score"),
        avg_grounding=_avg(circ_evals, "grounding_score"),
        avg_relevance=_avg(circ_evals, "relevance_score"),
        avg_accuracy=_avg(circ_evals, "accuracy_score"),
        avg_completeness=_avg(circ_evals, "completeness_score"),
    )
    return _call_claude(client, prompt)


# ── Save Results ──────────────────────────────────────────────────────────────

def _get_credentials():
    """Service-account credentials matching generate_tolerability_report.py pattern."""
    if CREDENTIALS_PATH and os.path.exists(CREDENTIALS_PATH):
        from google.oauth2 import service_account
        return service_account.Credentials.from_service_account_file(CREDENTIALS_PATH)
    return None


def upload_to_gcs(local_path: str) -> str:
    """Upload a file to GCS and return the gs:// URI."""
    from google.cloud import storage as gcs_storage

    credentials = _get_credentials()
    client = gcs_storage.Client(project=BQ_PROJECT_ID, credentials=credentials)
    bucket = client.bucket(GCS_BUCKET)

    filename = os.path.basename(local_path)
    blob_name = f"{GCS_BASE_PATH}/{filename}"
    gcs_uri = f"gs://{GCS_BUCKET}/{blob_name}"

    print(f"[GCS] Uploading → {gcs_uri}")
    bucket.blob(blob_name).upload_from_filename(
        local_path,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return gcs_uri


def save_eval_to_excel(rows, drug=None):
    """Build a clean, presentable multi-sheet comparison Excel and upload to GCS."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = os.path.join(OUTPUT_DIR, f"ipd3_eval_{drug or 'all'}_{ts}.xlsx")

    wb = Workbook()

    # ── Styles ────────────────────────────────────────────────────────────
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    title_font = Font(name="Arial", bold=True, size=14, color="1F3864")
    sub_font = Font(name="Arial", bold=True, size=11, color="2F5496")
    cell_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )
    green_fill = PatternFill("solid", fgColor="E2EFDA")
    red_fill = PatternFill("solid", fgColor="FCE4EC")
    yellow_fill = PatternFill("solid", fgColor="FFF9C4")
    light_gray = PatternFill("solid", fgColor="F5F5F5")

    def _score_fill(val):
        try:
            v = int(float(val))
        except (ValueError, TypeError):
            return None
        if v >= 4: return green_fill
        if v == 3: return yellow_fill
        if v <= 2: return red_fill
        return None

    def _write_header(ws, row, headers, col_start=1):
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=col_start + i, value=h)
            c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, center, thin

    def _write_row(ws, row, values, col_start=1, fonts=None, fills=None):
        for i, v in enumerate(values):
            c = ws.cell(row=row, column=col_start + i, value=v)
            c.font = (fonts[i] if fonts and i < len(fonts) else cell_font)
            c.alignment = center if i > 0 else left
            c.border = thin
            if fills and i < len(fills) and fills[i]:
                c.fill = fills[i]

    def _auto_width(ws, min_w=10, max_w=40):
        for col_cells in ws.columns:
            length = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
                min(max(length + 2, min_w), max_w)

    # ── Split rows by eval_type ───────────────────────────────────────────
    circ_rows = [r for r in rows if r.get("eval_type") == "circumvention"]
    synth_rows = [r for r in rows if r.get("eval_type") == "overall_synthesis"]

    agreed_count = sum(1 for r in circ_rows if str(r.get("eval_agreement")).lower() == "true")
    total_count = len(circ_rows) or 1

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 1: Summary (Gemini evaluated against Claude ground truth)
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="IPD3 Evaluation — Gemini vs Claude (ground truth)").font = title_font
    ws.cell(row=2, column=1, value=f"Drug: {drug or 'All'}  |  Date: {ts}  |  Judge: {CLAUDE_MODEL}").font = sub_font
    ws.cell(row=3, column=1, value="Claude Sonnet 4.6 output is treated as ground truth. Gemini is scored against it.").font = cell_font
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")
    ws.merge_cells("A3:D3")

    r = 5
    _write_header(ws, r, ["Metric", "Gemini Score (avg)", "Agreement Rate"])
    metrics = [
        ("Categories Evaluated", total_count, f"{agreed_count}/{total_count}"),
        ("Agreement Rate", f"{round(agreed_count/total_count*100,1)}%", ""),
        ("Avg Faithfulness", _avg_score(circ_rows, "eval_faithfulness_score"), ""),
        ("Avg Grounding", _avg_score(circ_rows, "eval_grounding_score"), ""),
        ("Avg Relevance", _avg_score(circ_rows, "eval_relevance_score"), ""),
        ("Avg Accuracy", _avg_score(circ_rows, "eval_accuracy_score"), ""),
        ("Avg Completeness", _avg_score(circ_rows, "eval_completeness_score"), ""),
        ("Avg Feasibility", _avg_score(circ_rows, "eval_feasibility_score"), ""),
        ("Avg Regulatory", _avg_score(circ_rows, "eval_regulatory_score"), ""),
    ]
    for i, (metric, val, extra) in enumerate(metrics):
        row_num = r + 1 + i
        fill_v = None
        if isinstance(val, (int, float)):
            fill_v = green_fill if val >= 4 else yellow_fill if val >= 3 else red_fill if val >= 1 else None
        _write_row(ws, row_num, [metric, val, extra],
                   fonts=[bold_font, cell_font, cell_font],
                   fills=[light_gray, fill_v, None])

    if synth_rows:
        synth_start = r + len(metrics) + 3
        ws.cell(row=synth_start, column=1, value="Per-Drug Verdict").font = sub_font
        for si, sr in enumerate(synth_rows):
            row_num = synth_start + 1 + si
            ws.cell(row=row_num, column=1, value=sr.get("Drug_Name", "")).font = bold_font
            ws.cell(row=row_num, column=2, value=str(sr.get("synth_recommendation", ""))).font = cell_font
            ws.cell(row=row_num, column=2).alignment = left
            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=4)

    _auto_width(ws)

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 2: Circumvention Comparison
    # ══════════════════════════════════════════════════════════════════════
    if circ_rows:
        ws2 = wb.create_sheet("Circumvention")
        headers = [
            "Drug", "Category",
            "Ground Truth Strategy\n(Claude)", "Gemini Strategy",
            "Agreement",
            "Faith", "Ground", "Relev",
            "Accur", "Compl", "Feasib", "Regul",
            "Deviation Explanation",
        ]
        _write_header(ws2, 1, headers)

        for i, r_data in enumerate(circ_rows):
            rn = i + 2
            agreed = str(r_data.get("eval_agreement", "")).lower() == "true"
            faith = r_data.get("eval_faithfulness_score")
            gnd = r_data.get("eval_grounding_score")
            rel = r_data.get("eval_relevance_score")
            acc = r_data.get("eval_accuracy_score")
            comp = r_data.get("eval_completeness_score")
            feas = r_data.get("eval_feasibility_score")
            reg = r_data.get("eval_regulatory_score")

            vals = [
                r_data.get("Drug_Name", ""),
                r_data.get("Patent_Category", ""),
                str(r_data.get("claude_Strategy") or ""),
                str(r_data.get("gemini_Strategy") or ""),
                "TRUE" if agreed else "FALSE",
                faith, gnd, rel, acc, comp, feas, reg,
                str(r_data.get("eval_deviation_explanation") or ""),
            ]
            fills = [
                None, None, None, None,
                green_fill if agreed else red_fill,
                _score_fill(faith), _score_fill(gnd), _score_fill(rel),
                _score_fill(acc), _score_fill(comp), _score_fill(feas), _score_fill(reg),
                None,
            ]
            _write_row(ws2, rn, vals, fills=fills)

        _auto_width(ws2, min_w=8, max_w=30)
        ws2.column_dimensions["C"].width = 55
        ws2.column_dimensions["D"].width = 55
        ws2.column_dimensions["M"].width = 65

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 3: Judge Notes
    # ══════════════════════════════════════════════════════════════════════
    if circ_rows:
        ws3 = wb.create_sheet("Judge Notes")
        headers = [
            "Drug", "Category", "Agreement",
            "Faithfulness Notes", "Grounding Notes",
            "Relevance Notes", "Overall Assessment",
        ]
        _write_header(ws3, 1, headers)

        for i, r_data in enumerate(circ_rows):
            rn = i + 2
            agreed = str(r_data.get("eval_agreement", "")).lower() == "true"
            vals = [
                r_data.get("Drug_Name", ""),
                r_data.get("Patent_Category", ""),
                "TRUE" if agreed else "FALSE",
                str(r_data.get("eval_faithfulness_notes", "")),
                str(r_data.get("eval_grounding_notes", "")),
                str(r_data.get("eval_relevance_notes", "")),
                str(r_data.get("eval_overall_assessment", "")),
            ]
            fills = [None, None, green_fill if agreed else red_fill, None, None, None, None]
            _write_row(ws3, rn, vals, fills=fills)

        _auto_width(ws3, min_w=12, max_w=65)

    # ── Save & upload ─────────────────────────────────────────────────────
    wb.save(fname)
    print(f"[EXCEL] Eval results → {fname}")

    try:
        gcs_uri = upload_to_gcs(fname)
        print(f"[GCS] {gcs_uri}")
    except Exception as e:
        print(f"[WARN] GCS upload failed: {e} — file saved locally at {fname}")

    return fname


def save_eval_to_bq(rows):
    if not rows: return
    from google.cloud import bigquery
    credentials = _get_credentials()
    client = bigquery.Client(project=BQ_PROJECT_ID, credentials=credentials, location=BQ_LOCATION)
    table_ref = f"{BQ_PROJECT_ID}.{BQ_DATASET_ID}.{EVAL_TABLE}"
    df = pd.DataFrame(rows)
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).replace({"None": None, "nan": None, "NaN": None})
    df["created_at"] = pd.Timestamp.now(tz="UTC")
    job_config = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_APPEND, autodetect=True,
        schema_update_options=[bigquery.SchemaUpdateOption.ALLOW_FIELD_ADDITION],
    )
    client.load_table_from_dataframe(df, table_ref, job_config=job_config, location=BQ_LOCATION).result()
    print(f"[BQ] {EVAL_TABLE}: {len(df)} rows written")


# ── Main Orchestrator ─────────────────────────────────────────────────────────

def orchestrate(drug=None, skip_run=False, skip_gemini=False, skip_claude=False,
                to_bq=False, extra_args=None, no_bq_pipelines=False):
    start = time.time()

    print("=" * 70)
    print("IPD3 ORCHESTRATOR — Gemini vs Claude Patent Thicket Evaluation")
    print(f"  Drug filter : {drug or 'ALL'}")
    print(f"  Judge model : {CLAUDE_MODEL}")
    print(f"  Eval output : {'BQ + Excel' if to_bq else 'Excel only'}")
    print(f"  JSON dir    : {OUTPUT_DIR}")
    print("=" * 70)

    # ── Step 1: Run both pipelines in parallel ─────────────────────────
    if not skip_run:
        pipeline_extra = list(extra_args or [])
        if no_bq_pipelines:
            pipeline_extra.append("--no-bq")

        run_pipelines_parallel(drug, pipeline_extra, skip_gemini, skip_claude)
    else:
        print("\n⏭️  Both pipelines skipped (--skip-run)")

    # ── Step 3: Load results from local JSON ─────────────────────────────
    print(f"\n{'='*60}")
    print("STEP 3: Loading results from local JSON files")
    print(f"  Directory: {OUTPUT_DIR}")
    print(f"{'='*60}")

    circ_df = load_circumvention_results(drug)

    if circ_df.empty:
        print("\n[ERROR] No circumvention results found. Ensure pipelines ran successfully and")
        print(f"  JSON files exist in {OUTPUT_DIR}/")
        print(f"  Expected: circumvention_gemini.json, circumvention_claude.json")
        return {"error": "No data", "output_rows": []}

    # ── Step 4: LLM-as-Judge evaluation ──────────────────────────────────
    print(f"\n{'='*60}")
    print("STEP 4: LLM-as-Judge — Gemini vs Claude (ground truth)")
    print(f"  Circumvention rows: {len(circ_df)}")
    print(f"  Ground truth: Claude Sonnet 4.6")
    print(f"{'='*60}")

    client = get_claude_client()
    eval_timestamp = datetime.now().isoformat()
    all_output_rows = []
    circ_evals = []

    # ── Evaluate circumvention (Gemini against Claude ground truth) ───────
    circ_rows = circ_df.to_dict("records") if not circ_df.empty else []
    if circ_rows:
        print(f"\nEvaluating {len(circ_rows)} categories (Gemini vs ground truth) ...")

        def _eval_circ(row):
            try:
                dn = row.get("Drug_Name", "?")
                cat = row.get("Patent_Category", "?")
                print(f"  [{dn}/{cat}] Evaluating ...")
                result = evaluate_circumvention_row(client, row)
                circ_evals.append(result)
                out = dict(row)
                out["eval_timestamp"] = eval_timestamp
                out["eval_type"] = "circumvention"
                for k, v in result.items():
                    out[f"eval_{k}"] = "; ".join(str(x) for x in v) if isinstance(v, list) else v
                return out
            except Exception as e:
                print(f"  [ERROR] {row.get('Drug_Name')}/{row.get('Patent_Category')}: {e}")
                out = dict(row)
                out["eval_timestamp"] = eval_timestamp
                out["eval_error"] = str(e)
                return out

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(_eval_circ, r): r for r in circ_rows}
            for f in as_completed(futures):
                out = f.result()
                if out: all_output_rows.append(out)

    # ── Overall synthesis per drug ────────────────────────────────────────
    drugs_evaluated = {r.get("Drug_Name") for r in all_output_rows if r.get("Drug_Name")}
    for d in sorted(drugs_evaluated):
        print(f"\n  Generating synthesis for {d} ...")
        d_circ = [e for e, r in zip(circ_evals, circ_rows) if r.get("Drug_Name") == d] if circ_rows else []
        try:
            synth = run_overall_synthesis(client, d, d_circ)
            all_output_rows.append({
                "Drug_Name": d, "eval_type": "overall_synthesis",
                "eval_timestamp": eval_timestamp,
                **{f"synth_{k}": ("; ".join(v) if isinstance(v, list) else v) for k, v in synth.items()},
            })
        except Exception as e:
            print(f"  [ERROR] Synthesis for {d}: {e}")

    # ── Summary ───────────────────────────────────────────────────────────
    agreed = sum(1 for e in circ_evals if e.get("agreement"))
    total = len(circ_evals) or 1

    def _avg(evals, key):
        vals = [e.get(key, 0) for e in evals if e.get(key)]
        return round(sum(vals) / len(vals), 1) if vals else 0

    print(f"\n{'='*60}")
    print("EVALUATION SUMMARY (Gemini scored against Claude ground truth)")
    print(f"  Categories evaluated     : {len(circ_evals)}")
    print(f"  Gemini agrees with GT    : {agreed}/{total} ({round(agreed/total*100,1)}%)")
    print(f"  Avg Faithfulness         : {_avg(circ_evals, 'faithfulness_score')}")
    print(f"  Avg Grounding            : {_avg(circ_evals, 'grounding_score')}")
    print(f"  Avg Relevance            : {_avg(circ_evals, 'relevance_score')}")
    print(f"  Avg Accuracy             : {_avg(circ_evals, 'accuracy_score')}")
    print(f"  Avg Completeness         : {_avg(circ_evals, 'completeness_score')}")
    print(f"  Drugs evaluated          : {len(drugs_evaluated)}")
    print(f"  Elapsed                  : {time.time()-start:.1f}s")
    print("=" * 60)

    # ── Save ──────────────────────────────────────────────────────────────
    save_eval_to_excel(all_output_rows, drug)
    if to_bq:
        try:
            save_eval_to_bq(all_output_rows)
        except Exception as e:
            print(f"[BQ] Write failed ({e}) — results already saved to Excel")

    return {
        "output_rows": all_output_rows,
        "summary": {
            "agreed": agreed, "total": total,
            "agreement_pct": round(agreed / total * 100, 1),
            "drugs_evaluated": len(drugs_evaluated),
        },
    }


# ── CLI ────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="IPD3 Orchestrator — Run Gemini & Claude, Then LLM-as-Judge Evaluation"
    )
    parser.add_argument("--drug", default=None, help="Drug name (omit = all)")
    parser.add_argument("--to-bq", action="store_true", help="Also write eval to BigQuery")
    parser.add_argument("--no-bq", action="store_true",
                        help="Pass --no-bq to sub-pipelines (skip BQ writes in pipelines)")
    parser.add_argument("--skip-run", action="store_true", help="Skip pipelines, eval only")
    parser.add_argument("--skip-gemini", action="store_true")
    parser.add_argument("--skip-claude", action="store_true")
    parser.add_argument("--skip-circumvention", action="store_true")
    parser.add_argument("--refresh-scores", action="store_true")
    parser.add_argument("--rerun", action="store_true")
    parser.add_argument("--csv-input", default=None,
                        help="Path to local CSV/Excel export of Master_LOE table. "
                             "Passed to both pipelines so they skip BigQuery reads.")
    args = parser.parse_args()

    extra = []
    if args.skip_circumvention: extra.append("--skip-circumvention")
    if args.refresh_scores: extra.append("--refresh-scores")
    if args.rerun: extra.append("--rerun")
    if args.csv_input:
        extra.extend(["--csv-input", args.csv_input])

    result = orchestrate(
        drug=args.drug, skip_run=args.skip_run,
        skip_gemini=args.skip_gemini, skip_claude=args.skip_claude,
        to_bq=args.to_bq, extra_args=extra or None,
        no_bq_pipelines=args.no_bq,
    )
    if "error" in result:
        print(f"Error: {result['error']}")
        sys.exit(1)
