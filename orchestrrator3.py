"""
IPD3 Orchestrator v3 — Strategy-Level Evaluation (per-Gemini-strategy)

Key improvement over orchestrator2.py:
  orchestrator2.py grouped ALL strategies per drug+category and sent the
  complete Gemini set vs. the complete Claude (ground truth) set to the judge
  in ONE call per category. That means an individual Gemini strategy's
  correctness gets diluted into a single category-wide verdict.

  This file instead takes EACH Gemini strategy on its own and compares it
  against the FULL set of Claude ground-truth strategies for that category,
  in its own judge call. So a category with 3 Gemini strategies produces 3
  separate judge calls (each strategy verified individually against all N
  ground-truth strategies), rather than 1 call for the whole category.

  Claude output = ground truth. Each Gemini strategy is scored against it.

Usage:
    python3 orchestrator3.py                           # run all
    python3 orchestrator3.py --drug Semaglutide         # single drug
    python3 orchestrator3.py --skip-run                 # eval only
    python3 orchestrator3.py --rerun                    # bypass checkpoint
"""

import os
import re
import sys
import json
import time
import random
import argparse
import subprocess
from datetime import datetime
from typing import Any, Dict, List, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
from anthropic import AnthropicVertex, RateLimitError

try:
    from dotenv import load_dotenv
    load_dotenv(override=True)
except ImportError:
    pass

try:
    from json_repair import repair_json as _repair_json_lib
except ImportError:
    _repair_json_lib = None

# ── Config ─────────────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

BQ_PROJECT_ID = os.getenv("BQ_PROJECT_ID", "cognito-prod-394707")
BQ_DATASET_ID = os.getenv("BQ_DATASET_ID", "cognito_prod_datamart")
BQ_LOCATION   = os.getenv("BQ_LOCATION", "asia-south1")
GCP_REGION    = os.getenv("GCP_REGION", "us-east5")
PROJECT_ID    = os.getenv("BQ_PROJECT_ID", "cognito-prod-394707")

CREDENTIALS_PATH = os.environ.get("CREDENTIALS_PATH") or os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "")
if CREDENTIALS_PATH and not os.path.isabs(CREDENTIALS_PATH):
    CREDENTIALS_PATH = os.path.join(SCRIPT_DIR, CREDENTIALS_PATH)

GCS_BUCKET    = os.getenv("GCS_BUCKET", "cognito-gcs")
GCS_BASE_PATH = os.getenv("GCS_EVAL_PATH", "Cognito_new/eval_reports")

CLAUDE_MODEL   = os.getenv("CLAUDE_MODEL", "claude-sonnet-4-6")
MAX_WORKERS    = 6
LLM_MAX_TOKENS = 16384

OUTPUT_DIR = os.getenv("IPD3_OUTPUT_DIR", os.path.join(SCRIPT_DIR, "ipd3_output"))
EVAL_TABLE = "IPD3_Eval_Table_v3"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_credentials():
    if CREDENTIALS_PATH and os.path.exists(CREDENTIALS_PATH):
        from google.oauth2 import service_account
        return service_account.Credentials.from_service_account_file(CREDENTIALS_PATH)
    return None


def get_claude_client() -> AnthropicVertex:
    return AnthropicVertex(region=GCP_REGION, project_id=PROJECT_ID)


def _call_claude(client: AnthropicVertex, prompt: str, max_retries: int = 3) -> Dict[str, Any]:
    for attempt in range(max_retries + 1):
        try:
            resp = client.messages.create(
                model=CLAUDE_MODEL, max_tokens=LLM_MAX_TOKENS,
                messages=[{"role": "user", "content": prompt}],
            )
            text = resp.content[0].text.strip()
            text = re.sub(r"^```(?:json)?", "", text).strip()
            text = re.sub(r"```$", "", text).strip()
            try:
                return json.loads(text)
            except json.JSONDecodeError:
                if _repair_json_lib:
                    return json.loads(_repair_json_lib(text))
                text = re.sub(r",\s*}", "}", text)
                text = re.sub(r",\s*]", "]", text)
                return json.loads(text)
        except RateLimitError:
            if attempt == max_retries: raise
            time.sleep(2.0 * (2 ** attempt) * (1 + random.uniform(0, 0.25)))
        except Exception as e:
            if attempt == max_retries:
                return {"error": str(e)}
            time.sleep(2.0 * (2 ** attempt))
    return {}


def _find_script(name):
    for c in [name, name.replace("(", "_").replace(")", "_")]:
        p = os.path.join(SCRIPT_DIR, c)
        if os.path.exists(p):
            return p
    return os.path.join(SCRIPT_DIR, name)


def _load_json_safe(path: str) -> List[Dict]:
    if not os.path.exists(path):
        print(f"[WARN] Not found: {path}")
        return []
    try:
        with open(path) as f:
            data = json.load(f)
        print(f"[JSON] Loaded {len(data)} rows from {path}")
        return data
    except Exception as e:
        print(f"[WARN] Failed: {path}: {e}")
        return []


# ── Run Pipelines (parallel) ─────────────────────────────────────────────────

def run_pipelines_parallel(drug=None, extra_args=None,
                           skip_gemini=False, skip_claude=False):
    procs = {}
    for name, script in [("Gemini", "ipd3bq__6_.py"), ("Claude", "ipd3bq_claude.py")]:
        if (name == "Gemini" and skip_gemini) or (name == "Claude" and skip_claude):
            print(f"\n⏭️  {name} pipeline skipped")
            continue
        cmd = [sys.executable, _find_script(script)]
        if drug: cmd.append(drug)
        if extra_args: cmd.extend(extra_args)
        print(f"\n  Launching {name}: {' '.join(cmd)}")
        procs[name] = subprocess.Popen(cmd)

    for name, proc in procs.items():
        rc = proc.wait()
        status = "Done" if rc == 0 else f"exited with code {rc}"
        print(f"  [{name}] {status}")


# ── Load & Group by Category ─────────────────────────────────────────────────

def load_and_group(drug=None) -> List[Dict]:
    """
    Load both JSONs, group ALL strategies per (Drug, Category) into a single
    record. Returns list of dicts, one per category, each with:
        Drug_Name, Patent_Category, Patents,
        gemini_strategies (list of dicts), claude_strategies (list of dicts),
        gemini_difficulty, claude_difficulty, gemini_summary, claude_summary
    """
    gemini_rows = _load_json_safe(os.path.join(OUTPUT_DIR, "circumvention_gemini.json"))
    claude_rows = _load_json_safe(os.path.join(OUTPUT_DIR, "circumvention_claude.json"))

    if not gemini_rows and not claude_rows:
        return []

    def _group(rows):
        grouped = {}
        for r in rows:
            key = (r.get("Drug_Name", "").strip(), r.get("Patent_Category", "").strip())
            if key not in grouped:
                grouped[key] = {
                    "strategies": [],
                    "difficulty": r.get("Overall_Difficulty", ""),
                    "summary": r.get("Summary", ""),
                    "patents": r.get("Patents", ""),
                    "key_claims": r.get("Key_Claim_Limitations", ""),
                    "fda_precedents": r.get("FDA_Precedents", ""),
                }
            strat = r.get("Strategy", "")
            if strat and strat != "No strategies identified":
                grouped[key]["strategies"].append({
                    "strategy": strat,
                    "rationale": r.get("Rationale", ""),
                    "feasibility": r.get("Feasibility", ""),
                    "regulatory_pathway": r.get("Regulatory_Pathway", ""),
                    "prior_art_support": r.get("Prior_Art_Support", ""),
                })
        return grouped

    g_grouped = _group(gemini_rows)
    c_grouped = _group(claude_rows)

    all_keys = set(g_grouped.keys()) | set(c_grouped.keys())
    if drug:
        all_keys = {k for k in all_keys if k[0].lower() == drug.lower()}

    results = []
    for (d, cat) in sorted(all_keys):
        g = g_grouped.get((d, cat), {"strategies": [], "difficulty": "", "summary": "", "patents": "", "key_claims": "", "fda_precedents": ""})
        c = c_grouped.get((d, cat), {"strategies": [], "difficulty": "", "summary": "", "patents": "", "key_claims": "", "fda_precedents": ""})
        results.append({
            "Drug_Name": d,
            "Patent_Category": cat,
            "Patents": g["patents"] or c["patents"],
            "gemini_strategies": g["strategies"],
            "claude_strategies": c["strategies"],
            "gemini_difficulty": g["difficulty"],
            "claude_difficulty": c["difficulty"],
            "gemini_summary": g["summary"],
            "claude_summary": c["summary"],
            "gemini_key_claims": g["key_claims"],
            "claude_key_claims": c["key_claims"],
            "gemini_fda_precedents": g["fda_precedents"],
            "claude_fda_precedents": c["fda_precedents"],
            "gemini_strategy_count": len(g["strategies"]),
            "claude_strategy_count": len(c["strategies"]),
        })

    print(f"[GROUP] {len(results)} categories (from {len(gemini_rows)} Gemini + {len(claude_rows)} Claude rows)")
    return results


# ── Strategy-Level Eval Prompt ────────────────────────────────────────────────

def _format_strategies(strategies: list) -> str:
    if not strategies:
        return "  (no strategies identified)"
    lines = []
    for i, s in enumerate(strategies, 1):
        lines.append(f"  Strategy {i}: {s.get('strategy', '')}")
        lines.append(f"    Rationale: {s.get('rationale', '')}")
        lines.append(f"    Feasibility: {s.get('feasibility', '')}")
        lines.append(f"    Regulatory: {s.get('regulatory_pathway', '')}")
        lines.append(f"    Prior Art: {s.get('prior_art_support', '')}")
    return "\n".join(lines)


# Each Gemini strategy is verified INDIVIDUALLY against the full ground-truth
# strategy set — one judge call per Gemini strategy, not one call per category.
STRATEGY_EVAL_PROMPT = """
You are a senior pharmaceutical patent attorney evaluating a single circumvention
strategy proposed by Gemini 2.5 Flash.

Claude Sonnet 4.6's complete strategy set for this category is the GROUND TRUTH.
You are evaluating ONE specific Gemini strategy against the FULL ground-truth set
below to determine whether it is correct, grounded, and consistent with an
established or plausible approach.

Drug: {drug}
Patent Category: {category}
Patents: {patents}

══ GROUND TRUTH (Claude — {gt_count} strategies) ══
Difficulty: {claude_difficulty}
Key Claims: {claude_key_claims}
{claude_strategies}
Summary: {claude_summary}
══ END GROUND TRUTH ══

══ GEMINI STRATEGY UNDER REVIEW (strategy {gemini_index} of {gemini_total} for this category) ══
{gemini_strategy}
══ END GEMINI STRATEGY ══

Evaluate THIS SINGLE Gemini strategy against the full ground-truth set:

1. **Agreement**: Does this strategy align with an approach found in (or clearly
   consistent with) the ground truth set? Consider whether it targets the same
   kind of claim vulnerability as one or more ground-truth strategies.

2. **Faithfulness**: Does it avoid hallucinated references or fabricated details?

3. **Grounding**: Is the claim traceable to the patent data provided?

4. **Relevance**: Is it specific to this drug/category (not generic boilerplate)?

5. **Accuracy**: Is it technically/legally sound when checked against the ground
   truth set?

6. **Completeness**: Does it address the relevant claim limitation(s) thoroughly?

7. **Feasibility**: Is its feasibility rating consistent with what the ground
   truth would imply for a similar approach?

8. **Regulatory Viability**: Is the 505(b)(2) pathway (if any) realistic?

Also identify:
- Which single ground-truth strategy (if any) this Gemini strategy most closely
  matches, and how strong that match is.

Respond ONLY with valid JSON:
{{
  "agreement": <true if this Gemini strategy substantially agrees with/maps to the ground truth, false otherwise>,
  "faithfulness_score": <1-5>,
  "grounding_score": <1-5>,
  "relevance_score": <1-5>,
  "accuracy_score": <1-5>,
  "completeness_score": <1-5>,
  "feasibility_score": <1-5>,
  "regulatory_score": <1-5>,
  "best_matched_gt_strategy": "<brief description of the closest-matching ground truth strategy, or null if none>",
  "match_quality": "<exact | partial | none>",
  "faithfulness_notes": "<1-2 sentences>",
  "grounding_notes": "<1-2 sentences>",
  "relevance_notes": "<1-2 sentences>",
  "overall_assessment": "<2-3 sentence verdict on this specific strategy>"
}}
"""

SYNTHESIS_PROMPT = """
You are a senior pharmaceutical patent attorney providing a final synthesis.

Claude Sonnet 4.6's output is the GROUND TRUTH. You evaluated each individual
Gemini strategy against the full ground-truth set for its category.

Drug: {drug}

Summary:
- Gemini strategies evaluated: {num_strategies} (across {num_categories} categories)
- Gemini strategies that agreed with ground truth: {agreed}/{num_strategies}
- Avg Faithfulness: {avg_faith} | Avg Grounding: {avg_gnd}
- Avg Relevance: {avg_rel} | Avg Accuracy: {avg_acc}
- Avg Completeness: {avg_comp}

Respond ONLY with valid JSON:
{{
  "gemini_overall_correct": <true/false>,
  "confidence": "<high|medium|low>",
  "key_strengths": ["<patterns where Gemini strategies matched ground truth>"],
  "key_weaknesses": ["<patterns where Gemini strategies deviated>"],
  "strategy_coverage": "<1-2 sentences on how well Gemini's strategies held up individually>",
  "recommendation": "<2-3 sentence verdict>"
}}
"""


# ── Evaluate One Gemini Strategy (against full GT set) ────────────────────────

def evaluate_strategy(client, cat_data: Dict, gemini_strategy: Dict, gemini_index: int) -> Dict:
    prompt = STRATEGY_EVAL_PROMPT.format(
        drug=cat_data["Drug_Name"],
        category=cat_data["Patent_Category"],
        patents=cat_data["Patents"],
        claude_difficulty=cat_data["claude_difficulty"],
        claude_key_claims=cat_data["claude_key_claims"],
        claude_strategies=_format_strategies(cat_data["claude_strategies"]),
        claude_summary=cat_data["claude_summary"],
        gt_count=cat_data["claude_strategy_count"],
        gemini_index=gemini_index,
        gemini_total=cat_data["gemini_strategy_count"],
        gemini_strategy=_format_strategies([gemini_strategy]),
    )
    return _call_claude(client, prompt)


def build_strategy_tasks(categories: List[Dict]) -> List[Dict]:
    """
    Flatten categories into one task per Gemini strategy. Each task still
    carries the full Claude (ground truth) strategy set for its category, so
    every Gemini strategy is checked against ALL ground-truth strategies.
    """
    tasks = []
    for cat_data in categories:
        gem_strats = cat_data["gemini_strategies"]
        if not gem_strats:
            # No Gemini strategies for this category — record a placeholder
            # so the category still shows up in the report.
            tasks.append({
                "Drug_Name": cat_data["Drug_Name"],
                "Patent_Category": cat_data["Patent_Category"],
                "gemini_index": 0,
                "gemini_total": 0,
                "gemini_strategy": None,
                "claude_strategy_count": cat_data["claude_strategy_count"],
                "cat_data": cat_data,
                "eval": {"skipped": True, "reason": "No Gemini strategies identified for this category"},
            })
            continue
        for idx, gs in enumerate(gem_strats, 1):
            tasks.append({
                "Drug_Name": cat_data["Drug_Name"],
                "Patent_Category": cat_data["Patent_Category"],
                "gemini_index": idx,
                "gemini_total": len(gem_strats),
                "gemini_strategy": gs,
                "claude_strategy_count": cat_data["claude_strategy_count"],
                "cat_data": cat_data,
            })
    return tasks


# ── GCS Upload ────────────────────────────────────────────────────────────────

def upload_to_gcs(local_path: str) -> str:
    from google.cloud import storage as gcs_storage
    credentials = _get_credentials()
    client = gcs_storage.Client(project=BQ_PROJECT_ID, credentials=credentials)
    bucket = client.bucket(GCS_BUCKET)
    blob_name = f"{GCS_BASE_PATH}/{os.path.basename(local_path)}"
    gcs_uri = f"gs://{GCS_BUCKET}/{blob_name}"
    print(f"[GCS] Uploading → {gcs_uri}")
    bucket.blob(blob_name).upload_from_filename(
        local_path,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return gcs_uri


# ── Excel Builder ─────────────────────────────────────────────────────────────

def save_eval_to_excel(strategy_results: List[Dict], category_agg: List[Dict],
                        synth_results: List[Dict], drug=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = os.path.join(OUTPUT_DIR, f"ipd3_eval_v3_{drug or 'all'}_{ts}.xlsx")
    wb = Workbook()

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    title_font = Font(name="Arial", bold=True, size=14, color="1F3864")
    sub_font = Font(name="Arial", bold=True, size=11, color="2F5496")
    cell_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    green = PatternFill("solid", fgColor="E2EFDA")
    red = PatternFill("solid", fgColor="FCE4EC")
    yellow = PatternFill("solid", fgColor="FFF9C4")
    gray = PatternFill("solid", fgColor="F5F5F5")

    def _sf(val):
        try: v = int(float(val))
        except: return None
        return green if v >= 4 else yellow if v == 3 else red if v <= 2 else None

    def _hdr(ws, row, headers):
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=1+i, value=h)
            c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, center, thin

    def _row(ws, row, values, fills=None):
        for i, v in enumerate(values):
            c = ws.cell(row=row, column=1+i, value=v)
            c.font, c.alignment, c.border = cell_font, (left_al if i <= 2 else center), thin
            if fills and i < len(fills) and fills[i]:
                c.fill = fills[i]

    def _auto(ws, mn=10, mx=40):
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+2, mn), mx)

    scored = [r for r in strategy_results if not r.get("eval", {}).get("skipped")]

    def _avg(key):
        vals = [r["eval"].get(key) for r in scored if r.get("eval", {}).get(key) is not None]
        nums = [float(v) for v in vals if v is not None]
        return round(sum(nums)/len(nums), 1) if nums else "—"

    total_strats = len(scored)
    agreed = sum(1 for r in scored if r.get("eval", {}).get("agreement"))
    total_categories = len(category_agg)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="IPD3 Eval v3 — Gemini vs Claude (ground truth), strategy-level").font = title_font
    ws.cell(row=2, column=1, value=f"Drug: {drug or 'All'}  |  Judge: {CLAUDE_MODEL}  |  Each Gemini strategy verified individually against all Claude strategies").font = sub_font
    ws.merge_cells("A1:D1"); ws.merge_cells("A2:D2")

    _hdr(ws, 4, ["Metric", "Value", ""])
    summary_data = [
        ("Categories Evaluated", total_categories, ""),
        ("Gemini Strategies Evaluated", total_strats, ""),
        ("Agreement Rate", f"{agreed}/{total_strats} ({round(agreed/(total_strats or 1)*100,1)}%)", ""),
        ("", "", ""),
        ("Avg Faithfulness", _avg("faithfulness_score"), ""),
        ("Avg Grounding", _avg("grounding_score"), ""),
        ("Avg Relevance", _avg("relevance_score"), ""),
        ("Avg Accuracy", _avg("accuracy_score"), ""),
        ("Avg Completeness", _avg("completeness_score"), ""),
        ("Avg Feasibility", _avg("feasibility_score"), ""),
        ("Avg Regulatory", _avg("regulatory_score"), ""),
    ]
    for i, (m, v, n) in enumerate(summary_data):
        fills = [gray, _sf(v) if isinstance(v, (int, float)) else None, None]
        _row(ws, 5+i, [m, v, n], fills=fills)
        if m: ws.cell(row=5+i, column=1).font = bold_font

    if synth_results:
        sr = len(summary_data) + 7
        ws.cell(row=sr, column=1, value="Per-Drug Verdict").font = sub_font
        for si, s in enumerate(synth_results):
            ws.cell(row=sr+1+si, column=1, value=s.get("drug", "")).font = bold_font
            ws.cell(row=sr+1+si, column=2, value=str(s.get("synth", {}).get("recommendation", ""))).font = cell_font
            ws.cell(row=sr+1+si, column=2).alignment = left_al
            ws.merge_cells(start_row=sr+1+si, start_column=2, end_row=sr+1+si, end_column=4)

    _auto(ws)

    # ── Sheet 2: Category Rollup ───────────────────────────────────────────
    ws2 = wb.create_sheet("Category Rollup")
    _hdr(ws2, 1, [
        "Drug", "Category", "GT Count", "Gemini Count",
        "Gemini Strategies Agreed", "Agreement Rate",
        "Avg Faith", "Avg Ground", "Avg Relev", "Avg Accur",
        "Avg Compl", "Avg Feasib", "Avg Regul",
    ])
    for i, r in enumerate(category_agg):
        rn = i + 2
        vals = [
            r["Drug_Name"], r["Patent_Category"],
            r["claude_strategy_count"], r["gemini_strategy_count"],
            f"{r['agreed']}/{r['scored']}" if r["scored"] else "—",
            f"{round(r['agree_rate']*100,1)}%" if r["scored"] else "—",
            r["avg_faith"], r["avg_ground"], r["avg_relev"], r["avg_accur"],
            r["avg_compl"], r["avg_feasib"], r["avg_regul"],
        ]
        fills = [None, None, None, None,
                 green if r["scored"] and r["agree_rate"] >= 0.5 else (red if r["scored"] else None),
                 None, None, None, None, None, None, None, None]
        _row(ws2, rn, vals, fills=fills)
    _auto(ws2)

    # ── Sheet 3: Strategy-Level Results (one row per Gemini strategy) ─────
    ws3 = wb.create_sheet("Strategy Results")
    _hdr(ws3, 1, [
        "Drug", "Category", "Gemini Strategy #",
        "Gemini Strategy", "Claude Strategy (all GT)", "Best-Matched GT Strategy", "Match Quality",
        "Agree", "Faith", "Ground", "Relev", "Accur", "Compl", "Feasib", "Regul",
        "Overall Assessment",
    ])
    for i, r in enumerate(strategy_results):
        rn = i + 2
        e = r.get("eval", {})
        claude_strats_all = "\n".join(
            f"• {s.get('strategy','')}" for s in r["cat_data"].get("claude_strategies", [])
        ) or "(none)"
        if e.get("skipped"):
            _row(ws3, rn, [
                r["Drug_Name"], r["Patent_Category"], "—",
                "(no Gemini strategies identified)", claude_strats_all, "", "",
                "N/A", "", "", "", "", "", "", "",
                e.get("reason", ""),
            ], fills=[None]*16)
            continue
        ag = e.get("agreement")
        gs = r.get("gemini_strategy") or {}
        vals = [
            r["Drug_Name"], r["Patent_Category"],
            f"{r['gemini_index']} of {r['gemini_total']}",
            gs.get("strategy", ""),
            claude_strats_all,
            str(e.get("best_matched_gt_strategy", "")),
            str(e.get("match_quality", "")),
            "TRUE" if ag else "FALSE",
            e.get("faithfulness_score"), e.get("grounding_score"),
            e.get("relevance_score"), e.get("accuracy_score"),
            e.get("completeness_score"), e.get("feasibility_score"),
            e.get("regulatory_score"),
            str(e.get("overall_assessment", "")),
        ]
        fills = [
            None, None, None, None, None, None, None,
            green if ag else red,
            _sf(e.get("faithfulness_score")), _sf(e.get("grounding_score")),
            _sf(e.get("relevance_score")), _sf(e.get("accuracy_score")),
            _sf(e.get("completeness_score")), _sf(e.get("feasibility_score")),
            _sf(e.get("regulatory_score")),
            None,
        ]
        _row(ws3, rn, vals, fills=fills)

    _auto(ws3, mn=8, mx=35)
    ws3.column_dimensions["D"].width = 55
    ws3.column_dimensions["E"].width = 55
    ws3.column_dimensions["F"].width = 45
    ws3.column_dimensions["P"].width = 55

    # ── Sheet 4: Judge Notes ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Judge Notes")
    _hdr(ws4, 1, [
        "Drug", "Category", "Gemini Strategy #", "Agreement",
        "Faithfulness Notes", "Grounding Notes", "Relevance Notes",
    ])
    ri = 2
    for r in strategy_results:
        e = r.get("eval", {})
        if e.get("skipped"):
            continue
        ag = e.get("agreement")
        _row(ws4, ri, [
            r["Drug_Name"], r["Patent_Category"], f"{r['gemini_index']} of {r['gemini_total']}",
            "TRUE" if ag else "FALSE",
            str(e.get("faithfulness_notes", "")),
            str(e.get("grounding_notes", "")),
            str(e.get("relevance_notes", "")),
        ], fills=[None, None, None, green if ag else red, None, None, None])
        ri += 1
    _auto(ws4, mn=12, mx=55)

    wb.save(fname)
    print(f"[EXCEL] → {fname}")
    try:
        gcs_uri = upload_to_gcs(fname)
        print(f"[GCS] {gcs_uri}")
    except Exception as e:
        print(f"[WARN] GCS upload failed: {e}")
    return fname


# ── Category-Level Aggregation (rolled up from strategy-level results) ────────

def aggregate_by_category(categories: List[Dict], strategy_results: List[Dict]) -> List[Dict]:
    by_cat = {}
    for r in strategy_results:
        key = (r["Drug_Name"], r["Patent_Category"])
        by_cat.setdefault(key, []).append(r)

    def _avg(evals, key):
        vals = [e.get(key) for e in evals if e.get(key) is not None]
        nums = [float(v) for v in vals if v is not None]
        return round(sum(nums)/len(nums), 1) if nums else "—"

    rollup = []
    for cat_data in categories:
        key = (cat_data["Drug_Name"], cat_data["Patent_Category"])
        rows = by_cat.get(key, [])
        scored_evals = [r["eval"] for r in rows if not r.get("eval", {}).get("skipped")]
        agreed = sum(1 for e in scored_evals if e.get("agreement"))
        n = len(scored_evals)
        rollup.append({
            "Drug_Name": cat_data["Drug_Name"],
            "Patent_Category": cat_data["Patent_Category"],
            "claude_strategy_count": cat_data["claude_strategy_count"],
            "gemini_strategy_count": cat_data["gemini_strategy_count"],
            "agreed": agreed,
            "scored": n,
            "agree_rate": (agreed / n) if n else 0,
            "avg_faith": _avg(scored_evals, "faithfulness_score"),
            "avg_ground": _avg(scored_evals, "grounding_score"),
            "avg_relev": _avg(scored_evals, "relevance_score"),
            "avg_accur": _avg(scored_evals, "accuracy_score"),
            "avg_compl": _avg(scored_evals, "completeness_score"),
            "avg_feasib": _avg(scored_evals, "feasibility_score"),
            "avg_regul": _avg(scored_evals, "regulatory_score"),
        })
    return rollup


# ── Main Orchestrator ─────────────────────────────────────────────────────────

def orchestrate(drug=None, skip_run=False, skip_gemini=False, skip_claude=False,
                no_bq_pipelines=False, extra_args=None):
    start = time.time()

    print("=" * 70)
    print("IPD3 ORCHESTRATOR v3 — Strategy-Level Evaluation")
    print(f"  Drug: {drug or 'ALL'}  |  Judge: {CLAUDE_MODEL}")
    print(f"  Ground truth: Claude Sonnet 4.6")
    print("  Each Gemini strategy is verified individually against ALL Claude strategies")
    print("=" * 70)

    if not skip_run:
        pipeline_extra = list(extra_args or [])
        if no_bq_pipelines:
            pipeline_extra.append("--no-bq")
        run_pipelines_parallel(drug, pipeline_extra, skip_gemini, skip_claude)
    else:
        print("\n⏭️  Pipelines skipped (--skip-run)")

    # ── Load & group ──────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print("Loading & grouping strategies by category")
    print(f"{'='*60}")
    categories = load_and_group(drug)

    if not categories:
        print("[ERROR] No data to evaluate.")
        return {"error": "No data"}

    # ── Build per-strategy tasks ─────────────────────────────────────────
    tasks = build_strategy_tasks(categories)
    to_score = [t for t in tasks if t.get("gemini_strategy") is not None]

    print(f"\n{'='*60}")
    print(f"Evaluating {len(to_score)} Gemini strategies across {len(categories)} categories")
    print("(1 judge call per Gemini strategy, each checked against its full GT set)")
    print(f"{'='*60}")

    client = get_claude_client()
    strategy_results = [t for t in tasks if t.get("gemini_strategy") is None]  # skipped placeholders

    def _eval_one(task):
        d, c = task["Drug_Name"], task["Patent_Category"]
        idx, tot = task["gemini_index"], task["gemini_total"]
        print(f"  [{d}/{c}] Gemini strategy {idx}/{tot} vs {task['claude_strategy_count']} GT strategies...")
        result = evaluate_strategy(client, task["cat_data"], task["gemini_strategy"], idx)
        return {**task, "eval": result}

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(_eval_one, t): t for t in to_score}
        for f in as_completed(futures):
            strategy_results.append(f.result())

    # Sort by drug, category, strategy index for consistent output
    strategy_results.sort(key=lambda r: (r["Drug_Name"], r["Patent_Category"], r["gemini_index"]))

    # ── Category rollup (for reporting) ─────────────────────────────────────
    category_agg = aggregate_by_category(categories, strategy_results)

    # ── Synthesis ─────────────────────────────────────────────────────────
    drugs = sorted({r["Drug_Name"] for r in strategy_results})
    synth_results = []
    for d in drugs:
        d_evals = [r["eval"] for r in strategy_results
                   if r["Drug_Name"] == d and not r.get("eval", {}).get("skipped")]
        d_cats = {r["Patent_Category"] for r in strategy_results if r["Drug_Name"] == d}
        _a = lambda k: round(sum(e.get(k, 0) for e in d_evals if e.get(k)) / (len(d_evals) or 1), 1)
        prompt = SYNTHESIS_PROMPT.format(
            drug=d, num_strategies=len(d_evals), num_categories=len(d_cats),
            agreed=sum(1 for e in d_evals if e.get("agreement")),
            avg_faith=_a("faithfulness_score"), avg_gnd=_a("grounding_score"),
            avg_rel=_a("relevance_score"), avg_acc=_a("accuracy_score"),
            avg_comp=_a("completeness_score"),
        )
        synth = _call_claude(client, prompt)
        synth_results.append({"drug": d, "synth": synth})

    # ── Summary ───────────────────────────────────────────────────────────
    scored = [r for r in strategy_results if not r.get("eval", {}).get("skipped")]
    agreed = sum(1 for r in scored if r.get("eval", {}).get("agreement"))
    total = len(scored)
    print(f"\n{'='*60}")
    print(f"SUMMARY — {agreed}/{total} Gemini strategies agreed with ground truth ({round(agreed/(total or 1)*100,1)}%)")
    print(f"  Elapsed: {time.time()-start:.1f}s")
    print("=" * 60)

    save_eval_to_excel(strategy_results, category_agg, synth_results, drug)
    return {"strategy_results": strategy_results, "category_agg": category_agg, "synth_results": synth_results}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="IPD3 Orchestrator v3 — Strategy-Level Eval")
    parser.add_argument("--drug", default=None)
    parser.add_argument("--no-bq", action="store_true")
    parser.add_argument("--skip-run", action="store_true")
    parser.add_argument("--skip-gemini", action="store_true")
    parser.add_argument("--skip-claude", action="store_true")
    parser.add_argument("--skip-circumvention", action="store_true")
    parser.add_argument("--refresh-scores", action="store_true")
    parser.add_argument("--rerun", action="store_true")
    parser.add_argument("--csv-input", default=None)
    args = parser.parse_args()

    extra = []
    if args.skip_circumvention: extra.append("--skip-circumvention")
    if args.refresh_scores: extra.append("--refresh-scores")
    if args.rerun: extra.append("--rerun")
    if args.csv_input: extra.extend(["--csv-input", args.csv_input])

    orchestrate(
        drug=args.drug, skip_run=args.skip_run,
        skip_gemini=args.skip_gemini, skip_claude=args.skip_claude,
        no_bq_pipelines=args.no_bq, extra_args=extra or None,
    )
