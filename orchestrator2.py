"""
IPD3 Orchestrator v2 — Category-Level Evaluation (not row-level)

Key improvement over orchestrator.py:
  The circumvention JSON has one ROW per strategy. orchestrator.py compared
  them row-by-row, creating an N×M cross-product. This file groups ALL
  strategies per drug+category and sends the complete set to the judge in
  ONE call. The judge sees the full picture for each category.

  Claude output = ground truth. Gemini is scored against it.

Usage:
    python3 orchestrator2.py                           # run all
    python3 orchestrator2.py --drug Semaglutide        # single drug
    python3 orchestrator2.py --skip-run                # eval only
    python3 orchestrator2.py --rerun                   # bypass checkpoint
"""

import os
import re
import sys
import json
import time
import random
import argparse
import subprocess
from datetime import datetime
from typing import Any, Dict, List, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
from anthropic import AnthropicVertex, RateLimitError

try:
    from dotenv import load_dotenv
    load_dotenv(override=True)
except ImportError:
    pass

try:
    from json_repair import repair_json as _repair_json_lib
except ImportError:
    _repair_json_lib = None

# ── Config ─────────────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

BQ_PROJECT_ID = os.getenv("BQ_PROJECT_ID", "cognito-prod-394707")
BQ_DATASET_ID = os.getenv("BQ_DATASET_ID", "cognito_prod_datamart")
BQ_LOCATION   = os.getenv("BQ_LOCATION", "asia-south1")
GCP_REGION    = os.getenv("GCP_REGION", "us-east5")
PROJECT_ID    = os.getenv("BQ_PROJECT_ID", "cognito-prod-394707")

CREDENTIALS_PATH = os.environ.get("CREDENTIALS_PATH") or os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "")
if CREDENTIALS_PATH and not os.path.isabs(CREDENTIALS_PATH):
    CREDENTIALS_PATH = os.path.join(SCRIPT_DIR, CREDENTIALS_PATH)

GCS_BUCKET    = os.getenv("GCS_BUCKET", "cognito-gcs")
GCS_BASE_PATH = os.getenv("GCS_EVAL_PATH", "Cognito_new/eval_reports")

CLAUDE_MODEL   = os.getenv("CLAUDE_MODEL", "claude-sonnet-4-6")
MAX_WORKERS    = 6
LLM_MAX_TOKENS = 16384

OUTPUT_DIR = os.getenv("IPD3_OUTPUT_DIR", os.path.join(SCRIPT_DIR, "ipd3_output"))
EVAL_TABLE = "IPD3_Eval_Table_v2"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_credentials():
    if CREDENTIALS_PATH and os.path.exists(CREDENTIALS_PATH):
        from google.oauth2 import service_account
        return service_account.Credentials.from_service_account_file(CREDENTIALS_PATH)
    return None


def get_claude_client() -> AnthropicVertex:
    return AnthropicVertex(region=GCP_REGION, project_id=PROJECT_ID)


def _call_claude(client: AnthropicVertex, prompt: str, max_retries: int = 3) -> Dict[str, Any]:
    for attempt in range(max_retries + 1):
        try:
            resp = client.messages.create(
                model=CLAUDE_MODEL, max_tokens=LLM_MAX_TOKENS,
                messages=[{"role": "user", "content": prompt}],
            )
            text = resp.content[0].text.strip()
            text = re.sub(r"^```(?:json)?", "", text).strip()
            text = re.sub(r"```$", "", text).strip()
            try:
                return json.loads(text)
            except json.JSONDecodeError:
                if _repair_json_lib:
                    return json.loads(_repair_json_lib(text))
                text = re.sub(r",\s*}", "}", text)
                text = re.sub(r",\s*]", "]", text)
                return json.loads(text)
        except RateLimitError:
            if attempt == max_retries: raise
            time.sleep(2.0 * (2 ** attempt) * (1 + random.uniform(0, 0.25)))
        except Exception as e:
            if attempt == max_retries:
                return {"error": str(e)}
            time.sleep(2.0 * (2 ** attempt))
    return {}


def _find_script(name):
    for c in [name, name.replace("(", "_").replace(")", "_")]:
        p = os.path.join(SCRIPT_DIR, c)
        if os.path.exists(p):
            return p
    return os.path.join(SCRIPT_DIR, name)


def _load_json_safe(path: str) -> List[Dict]:
    if not os.path.exists(path):
        print(f"[WARN] Not found: {path}")
        return []
    try:
        with open(path) as f:
            data = json.load(f)
        print(f"[JSON] Loaded {len(data)} rows from {path}")
        return data
    except Exception as e:
        print(f"[WARN] Failed: {path}: {e}")
        return []


# ── Run Pipelines (parallel) ─────────────────────────────────────────────────

def run_pipelines_parallel(drug=None, extra_args=None,
                           skip_gemini=False, skip_claude=False):
    procs = {}
    for name, script in [("Gemini", "ipd3bq__6_.py"), ("Claude", "ipd3bq_claude.py")]:
        if (name == "Gemini" and skip_gemini) or (name == "Claude" and skip_claude):
            print(f"\n⏭️  {name} pipeline skipped")
            continue
        cmd = [sys.executable, _find_script(script)]
        if drug: cmd.append(drug)
        if extra_args: cmd.extend(extra_args)
        print(f"\n  Launching {name}: {' '.join(cmd)}")
        procs[name] = subprocess.Popen(cmd)

    for name, proc in procs.items():
        rc = proc.wait()
        status = "Done" if rc == 0 else f"exited with code {rc}"
        print(f"  [{name}] {status}")


# ── Load & Group by Category ─────────────────────────────────────────────────

def load_and_group(drug=None) -> List[Dict]:
    """
    Load both JSONs, group ALL strategies per (Drug, Category) into a single
    record. Returns list of dicts, one per category, each with:
        Drug_Name, Patent_Category, Patents,
        gemini_strategies (list of dicts), claude_strategies (list of dicts),
        gemini_difficulty, claude_difficulty, gemini_summary, claude_summary
    """
    gemini_rows = _load_json_safe(os.path.join(OUTPUT_DIR, "circumvention_gemini.json"))
    claude_rows = _load_json_safe(os.path.join(OUTPUT_DIR, "circumvention_claude.json"))

    if not gemini_rows and not claude_rows:
        return []

    def _group(rows):
        grouped = {}
        for r in rows:
            key = (r.get("Drug_Name", "").strip(), r.get("Patent_Category", "").strip())
            if key not in grouped:
                grouped[key] = {
                    "strategies": [],
                    "difficulty": r.get("Overall_Difficulty", ""),
                    "summary": r.get("Summary", ""),
                    "patents": r.get("Patents", ""),
                    "key_claims": r.get("Key_Claim_Limitations", ""),
                    "fda_precedents": r.get("FDA_Precedents", ""),
                }
            strat = r.get("Strategy", "")
            if strat and strat != "No strategies identified":
                grouped[key]["strategies"].append({
                    "strategy": strat,
                    "rationale": r.get("Rationale", ""),
                    "feasibility": r.get("Feasibility", ""),
                    "regulatory_pathway": r.get("Regulatory_Pathway", ""),
                    "prior_art_support": r.get("Prior_Art_Support", ""),
                })
        return grouped

    g_grouped = _group(gemini_rows)
    c_grouped = _group(claude_rows)

    all_keys = set(g_grouped.keys()) | set(c_grouped.keys())
    if drug:
        all_keys = {k for k in all_keys if k[0].lower() == drug.lower()}

    results = []
    for (d, cat) in sorted(all_keys):
        g = g_grouped.get((d, cat), {"strategies": [], "difficulty": "", "summary": "", "patents": "", "key_claims": "", "fda_precedents": ""})
        c = c_grouped.get((d, cat), {"strategies": [], "difficulty": "", "summary": "", "patents": "", "key_claims": "", "fda_precedents": ""})
        results.append({
            "Drug_Name": d,
            "Patent_Category": cat,
            "Patents": g["patents"] or c["patents"],
            "gemini_strategies": g["strategies"],
            "claude_strategies": c["strategies"],
            "gemini_difficulty": g["difficulty"],
            "claude_difficulty": c["difficulty"],
            "gemini_summary": g["summary"],
            "claude_summary": c["summary"],
            "gemini_key_claims": g["key_claims"],
            "claude_key_claims": c["key_claims"],
            "gemini_fda_precedents": g["fda_precedents"],
            "claude_fda_precedents": c["fda_precedents"],
            "gemini_strategy_count": len(g["strategies"]),
            "claude_strategy_count": len(c["strategies"]),
        })

    print(f"[GROUP] {len(results)} categories (from {len(gemini_rows)} Gemini + {len(claude_rows)} Claude rows)")
    return results


# ── Category-Level Eval Prompt ────────────────────────────────────────────────

def _format_strategies(strategies: list) -> str:
    if not strategies:
        return "  (no strategies identified)"
    lines = []
    for i, s in enumerate(strategies, 1):
        lines.append(f"  Strategy {i}: {s.get('strategy', '')}")
        lines.append(f"    Rationale: {s.get('rationale', '')}")
        lines.append(f"    Feasibility: {s.get('feasibility', '')}")
        lines.append(f"    Regulatory: {s.get('regulatory_pathway', '')}")
        lines.append(f"    Prior Art: {s.get('prior_art_support', '')}")
    return "\n".join(lines)


CATEGORY_EVAL_PROMPT = """
You are a senior pharmaceutical patent attorney evaluating circumvention strategies.

Claude Sonnet 4.6's output is the GROUND TRUTH. You are evaluating whether
Gemini 2.5 Flash's complete set of strategies for this category is correct.

Drug: {drug}
Patent Category: {category}
Patents: {patents}

══ GROUND TRUTH (Claude — {gt_count} strategies) ══
Difficulty: {claude_difficulty}
Key Claims: {claude_key_claims}
{claude_strategies}
Summary: {claude_summary}
══ END GROUND TRUTH ══

══ GEMINI OUTPUT ({gemini_count} strategies) ══
Difficulty: {gemini_difficulty}
Key Claims: {gemini_key_claims}
{gemini_strategies}
Summary: {gemini_summary}
══ END GEMINI ══

Evaluate Gemini's COMPLETE strategy set against the ground truth:

1. **Agreement**: Does Gemini's overall analysis agree with the ground truth?
   Consider: Do they identify similar claim limitations? Do strategies target
   the same vulnerabilities? Is the difficulty assessment consistent?

2. **Faithfulness**: Does Gemini avoid hallucinated references or fabricated details?

3. **Grounding**: Are Gemini's claims traceable to patent data?

4. **Relevance**: Is Gemini's analysis specific to this drug/category?

5. **Accuracy**: Are strategies technically/legally sound per ground truth?

6. **Completeness**: Does Gemini cover all viable approaches from the ground truth?
   Note ground truth has {gt_count} strategies, Gemini has {gemini_count}.

7. **Feasibility**: Are Gemini's feasibility ratings consistent with ground truth?

8. **Regulatory Viability**: Are 505(b)(2) pathways realistic?

Also identify:
- Which Gemini strategies MATCH a ground truth strategy (even if worded differently)
- Which Gemini strategies have NO ground truth equivalent (novel or hallucinated)
- Which ground truth strategies are MISSING from Gemini's output

Respond ONLY with valid JSON:
{{
  "agreement": <true if Gemini substantially agrees with ground truth, false otherwise>,
  "faithfulness_score": <1-5>,
  "grounding_score": <1-5>,
  "relevance_score": <1-5>,
  "accuracy_score": <1-5>,
  "completeness_score": <1-5>,
  "feasibility_score": <1-5>,
  "regulatory_score": <1-5>,
  "matched_strategies": ["<brief description of each Gemini strategy that matches a GT strategy>"],
  "missing_from_gemini": ["<GT strategies not found in Gemini output>"],
  "novel_in_gemini": ["<Gemini strategies with no GT equivalent — may be valid or hallucinated>"],
  "difficulty_match": <true or false>,
  "faithfulness_notes": "<1-2 sentences>",
  "grounding_notes": "<1-2 sentences>",
  "relevance_notes": "<1-2 sentences>",
  "deviation_explanation": "<2-3 sentences on where Gemini deviates>",
  "overall_assessment": "<2-3 sentence verdict>"
}}
"""

SYNTHESIS_PROMPT = """
You are a senior pharmaceutical patent attorney providing a final synthesis.

Claude Sonnet 4.6's output is the GROUND TRUTH. You evaluated Gemini against it.

Drug: {drug}

Summary:
- Categories evaluated: {num_categories}
- Gemini agreed with ground truth: {agreed}/{num_categories}
- Avg Faithfulness: {avg_faith} | Avg Grounding: {avg_gnd}
- Avg Relevance: {avg_rel} | Avg Accuracy: {avg_acc}
- Avg Completeness: {avg_comp}
- Strategies matched: {total_matched} | Missing from Gemini: {total_missing} | Novel in Gemini: {total_novel}

Respond ONLY with valid JSON:
{{
  "gemini_overall_correct": <true/false>,
  "confidence": "<high|medium|low>",
  "key_strengths": ["<where Gemini matched ground truth>"],
  "key_weaknesses": ["<where Gemini deviated>"],
  "strategy_coverage": "<1-2 sentences on how well Gemini covered ground truth strategies>",
  "recommendation": "<2-3 sentence verdict>"
}}
"""


# ── Evaluate One Category ─────────────────────────────────────────────────────

def evaluate_category(client, cat_data):
    prompt = CATEGORY_EVAL_PROMPT.format(
        drug=cat_data["Drug_Name"],
        category=cat_data["Patent_Category"],
        patents=cat_data["Patents"],
        claude_difficulty=cat_data["claude_difficulty"],
        claude_key_claims=cat_data["claude_key_claims"],
        claude_strategies=_format_strategies(cat_data["claude_strategies"]),
        claude_summary=cat_data["claude_summary"],
        gt_count=cat_data["claude_strategy_count"],
        gemini_difficulty=cat_data["gemini_difficulty"],
        gemini_key_claims=cat_data["gemini_key_claims"],
        gemini_strategies=_format_strategies(cat_data["gemini_strategies"]),
        gemini_summary=cat_data["gemini_summary"],
        gemini_count=cat_data["gemini_strategy_count"],
    )
    return _call_claude(client, prompt)


# ── GCS Upload ────────────────────────────────────────────────────────────────

def upload_to_gcs(local_path: str) -> str:
    from google.cloud import storage as gcs_storage
    credentials = _get_credentials()
    client = gcs_storage.Client(project=BQ_PROJECT_ID, credentials=credentials)
    bucket = client.bucket(GCS_BUCKET)
    blob_name = f"{GCS_BASE_PATH}/{os.path.basename(local_path)}"
    gcs_uri = f"gs://{GCS_BUCKET}/{blob_name}"
    print(f"[GCS] Uploading → {gcs_uri}")
    bucket.blob(blob_name).upload_from_filename(
        local_path,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return gcs_uri


# ── Excel Builder ─────────────────────────────────────────────────────────────

def save_eval_to_excel(eval_results: List[Dict], synth_results: List[Dict], drug=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = os.path.join(OUTPUT_DIR, f"ipd3_eval_v2_{drug or 'all'}_{ts}.xlsx")
    wb = Workbook()

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    title_font = Font(name="Arial", bold=True, size=14, color="1F3864")
    sub_font = Font(name="Arial", bold=True, size=11, color="2F5496")
    cell_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    green = PatternFill("solid", fgColor="E2EFDA")
    red = PatternFill("solid", fgColor="FCE4EC")
    yellow = PatternFill("solid", fgColor="FFF9C4")
    gray = PatternFill("solid", fgColor="F5F5F5")

    def _sf(val):
        try: v = int(float(val))
        except: return None
        return green if v >= 4 else yellow if v == 3 else red if v <= 2 else None

    def _hdr(ws, row, headers):
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=1+i, value=h)
            c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, center, thin

    def _row(ws, row, values, fills=None):
        for i, v in enumerate(values):
            c = ws.cell(row=row, column=1+i, value=v)
            c.font, c.alignment, c.border = cell_font, (left_al if i <= 2 else center), thin
            if fills and i < len(fills) and fills[i]:
                c.fill = fills[i]

    def _auto(ws, mn=10, mx=40):
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+2, mn), mx)

    def _avg(key):
        vals = [r["eval"].get(key) for r in eval_results if r.get("eval", {}).get(key)]
        nums = [float(v) for v in vals if v is not None]
        return round(sum(nums)/len(nums), 1) if nums else "—"

    total = len(eval_results)
    agreed = sum(1 for r in eval_results if r.get("eval", {}).get("agreement"))
    total_matched = sum(len(r.get("eval", {}).get("matched_strategies", [])) for r in eval_results)
    total_missing = sum(len(r.get("eval", {}).get("missing_from_gemini", [])) for r in eval_results)
    total_novel = sum(len(r.get("eval", {}).get("novel_in_gemini", [])) for r in eval_results)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="IPD3 Eval v2 — Gemini vs Claude (ground truth)").font = title_font
    ws.cell(row=2, column=1, value=f"Drug: {drug or 'All'}  |  Judge: {CLAUDE_MODEL}  |  Category-level comparison").font = sub_font
    ws.merge_cells("A1:D1"); ws.merge_cells("A2:D2")

    _hdr(ws, 4, ["Metric", "Value", ""])
    summary_data = [
        ("Categories Evaluated", total, ""),
        ("Agreement Rate", f"{agreed}/{total} ({round(agreed/(total or 1)*100,1)}%)", ""),
        ("", "", ""),
        ("Avg Faithfulness", _avg("faithfulness_score"), ""),
        ("Avg Grounding", _avg("grounding_score"), ""),
        ("Avg Relevance", _avg("relevance_score"), ""),
        ("Avg Accuracy", _avg("accuracy_score"), ""),
        ("Avg Completeness", _avg("completeness_score"), ""),
        ("Avg Feasibility", _avg("feasibility_score"), ""),
        ("Avg Regulatory", _avg("regulatory_score"), ""),
        ("", "", ""),
        ("Strategies Matched", total_matched, "Gemini strategies that match GT"),
        ("Missing from Gemini", total_missing, "GT strategies Gemini missed"),
        ("Novel in Gemini", total_novel, "Gemini strategies with no GT match"),
    ]
    for i, (m, v, n) in enumerate(summary_data):
        fills = [gray, _sf(v) if isinstance(v, (int, float)) else None, None]
        _row(ws, 5+i, [m, v, n], fills=fills)
        if m: ws.cell(row=5+i, column=1).font = bold_font

    if synth_results:
        sr = len(summary_data) + 7
        ws.cell(row=sr, column=1, value="Per-Drug Verdict").font = sub_font
        for si, s in enumerate(synth_results):
            ws.cell(row=sr+1+si, column=1, value=s.get("drug", "")).font = bold_font
            ws.cell(row=sr+1+si, column=2, value=str(s.get("synth", {}).get("recommendation", ""))).font = cell_font
            ws.cell(row=sr+1+si, column=2).alignment = left_al
            ws.merge_cells(start_row=sr+1+si, start_column=2, end_row=sr+1+si, end_column=4)

    _auto(ws)

    # ── Sheet 2: Circumvention ────────────────────────────────────────────
    ws2 = wb.create_sheet("Circumvention")
    _hdr(ws2, 1, [
        "Drug", "Category",
        "GT Strategies\n(Claude)", "Gemini Strategies",
        "GT\nCount", "Gemini\nCount",
        "Agree", "Faith", "Ground", "Relev", "Accur", "Compl", "Feasib", "Regul",
        "Matched", "Missing from Gemini", "Novel in Gemini",
        "Deviation Explanation",
    ])

    for i, r in enumerate(eval_results):
        rn = i + 2
        e = r.get("eval", {})
        ag = e.get("agreement")
        gt_strats = "\n".join(f"• {s.get('strategy','')}" for s in r.get("claude_strategies", []))
        gm_strats = "\n".join(f"• {s.get('strategy','')}" for s in r.get("gemini_strategies", []))
        matched = "; ".join(e.get("matched_strategies", []))
        missing = "; ".join(e.get("missing_from_gemini", []))
        novel = "; ".join(e.get("novel_in_gemini", []))

        vals = [
            r["Drug_Name"], r["Patent_Category"],
            gt_strats or "(none)", gm_strats or "(none)",
            r["claude_strategy_count"], r["gemini_strategy_count"],
            "TRUE" if ag else "FALSE",
            e.get("faithfulness_score"), e.get("grounding_score"),
            e.get("relevance_score"), e.get("accuracy_score"),
            e.get("completeness_score"), e.get("feasibility_score"),
            e.get("regulatory_score"),
            matched, missing, novel,
            str(e.get("deviation_explanation", "")),
        ]
        fills = [
            None, None, None, None, None, None,
            green if ag else red,
            _sf(e.get("faithfulness_score")), _sf(e.get("grounding_score")),
            _sf(e.get("relevance_score")), _sf(e.get("accuracy_score")),
            _sf(e.get("completeness_score")), _sf(e.get("feasibility_score")),
            _sf(e.get("regulatory_score")),
            None, None, None, None,
        ]
        _row(ws2, rn, vals, fills=fills)

    _auto(ws2, mn=8, mx=30)
    ws2.column_dimensions["C"].width = 55
    ws2.column_dimensions["D"].width = 55
    ws2.column_dimensions["O"].width = 45
    ws2.column_dimensions["P"].width = 45
    ws2.column_dimensions["Q"].width = 45
    ws2.column_dimensions["R"].width = 60

    # ── Sheet 3: Judge Notes ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Judge Notes")
    _hdr(ws3, 1, [
        "Drug", "Category", "Agreement",
        "Faithfulness Notes", "Grounding Notes", "Relevance Notes",
        "Overall Assessment",
    ])
    for i, r in enumerate(eval_results):
        rn = i + 2
        e = r.get("eval", {})
        ag = e.get("agreement")
        _row(ws3, rn, [
            r["Drug_Name"], r["Patent_Category"],
            "TRUE" if ag else "FALSE",
            str(e.get("faithfulness_notes", "")),
            str(e.get("grounding_notes", "")),
            str(e.get("relevance_notes", "")),
            str(e.get("overall_assessment", "")),
        ], fills=[None, None, green if ag else red, None, None, None, None])
    _auto(ws3, mn=12, mx=65)

    wb.save(fname)
    print(f"[EXCEL] → {fname}")
    try:
        gcs_uri = upload_to_gcs(fname)
        print(f"[GCS] {gcs_uri}")
    except Exception as e:
        print(f"[WARN] GCS upload failed: {e}")
    return fname


# ── Main Orchestrator ─────────────────────────────────────────────────────────

def orchestrate(drug=None, skip_run=False, skip_gemini=False, skip_claude=False,
                no_bq_pipelines=False, extra_args=None):
    start = time.time()

    print("=" * 70)
    print("IPD3 ORCHESTRATOR v2 — Category-Level Evaluation")
    print(f"  Drug: {drug or 'ALL'}  |  Judge: {CLAUDE_MODEL}")
    print(f"  Ground truth: Claude Sonnet 4.6")
    print("=" * 70)

    if not skip_run:
        pipeline_extra = list(extra_args or [])
        if no_bq_pipelines:
            pipeline_extra.append("--no-bq")
        run_pipelines_parallel(drug, pipeline_extra, skip_gemini, skip_claude)
    else:
        print("\n⏭️  Pipelines skipped (--skip-run)")

    # ── Load & group ──────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print("Loading & grouping strategies by category")
    print(f"{'='*60}")
    categories = load_and_group(drug)

    if not categories:
        print("[ERROR] No data to evaluate.")
        return {"error": "No data"}

    # ── Evaluate ──────────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"Evaluating {len(categories)} categories (1 judge call per category)")
    print(f"{'='*60}")

    client = get_claude_client()
    eval_results = []

    def _eval_one(cat_data):
        d, c = cat_data["Drug_Name"], cat_data["Patent_Category"]
        print(f"  [{d}/{c}] {cat_data['gemini_strategy_count']}G vs {cat_data['claude_strategy_count']}C strategies...")
        result = evaluate_category(client, cat_data)
        return {**cat_data, "eval": result}

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(_eval_one, c): c for c in categories}
        for f in as_completed(futures):
            eval_results.append(f.result())

    # Sort by drug then category for consistent output
    eval_results.sort(key=lambda r: (r["Drug_Name"], r["Patent_Category"]))

    # ── Synthesis ─────────────────────────────────────────────────────────
    drugs = sorted({r["Drug_Name"] for r in eval_results})
    synth_results = []
    for d in drugs:
        d_evals = [r["eval"] for r in eval_results if r["Drug_Name"] == d]
        _a = lambda k: round(sum(e.get(k, 0) for e in d_evals if e.get(k)) / (len(d_evals) or 1), 1)
        prompt = SYNTHESIS_PROMPT.format(
            drug=d, num_categories=len(d_evals),
            agreed=sum(1 for e in d_evals if e.get("agreement")),
            avg_faith=_a("faithfulness_score"), avg_gnd=_a("grounding_score"),
            avg_rel=_a("relevance_score"), avg_acc=_a("accuracy_score"),
            avg_comp=_a("completeness_score"),
            total_matched=sum(len(e.get("matched_strategies", [])) for e in d_evals),
            total_missing=sum(len(e.get("missing_from_gemini", [])) for e in d_evals),
            total_novel=sum(len(e.get("novel_in_gemini", [])) for e in d_evals),
        )
        synth = _call_claude(client, prompt)
        synth_results.append({"drug": d, "synth": synth})

    # ── Summary ───────────────────────────────────────────────────────────
    agreed = sum(1 for r in eval_results if r.get("eval", {}).get("agreement"))
    total = len(eval_results)
    print(f"\n{'='*60}")
    print(f"SUMMARY — {agreed}/{total} categories agreed ({round(agreed/(total or 1)*100,1)}%)")
    print(f"  Elapsed: {time.time()-start:.1f}s")
    print("=" * 60)

    save_eval_to_excel(eval_results, synth_results, drug)
    return {"eval_results": eval_results, "synth_results": synth_results}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="IPD3 Orchestrator v2 — Category-Level Eval")
    parser.add_argument("--drug", default=None)
    parser.add_argument("--no-bq", action="store_true")
    parser.add_argument("--skip-run", action="store_true")
    parser.add_argument("--skip-gemini", action="store_true")
    parser.add_argument("--skip-claude", action="store_true")
    parser.add_argument("--skip-circumvention", action="store_true")
    parser.add_argument("--refresh-scores", action="store_true")
    parser.add_argument("--rerun", action="store_true")
    parser.add_argument("--csv-input", default=None)
    args = parser.parse_args()

    extra = []
    if args.skip_circumvention: extra.append("--skip-circumvention")
    if args.refresh_scores: extra.append("--refresh-scores")
    if args.rerun: extra.append("--rerun")
    if args.csv_input: extra.extend(["--csv-input", args.csv_input])

    orchestrate(
        drug=args.drug, skip_run=args.skip_run,
        skip_gemini=args.skip_gemini, skip_claude=args.skip_claude,
        no_bq_pipelines=args.no_bq, extra_args=extra or None,
    )
